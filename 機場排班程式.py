
import pandas as pd
import random
import os
import re
import copy
import statistics
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# =========================
# Debug switches
# NOTE: Keep DEBUG_SCHED=False by default to avoid console spam.
# Turn it on only when you are actively debugging scheduling behavior.
# =========================
DEBUG_SCHED = False
RESET_BEFORE_SCHED = True
SMART_TEAM_PICK = False  # fixed A/B by early/late rules; no dynamic switching

SHIFT_ORDER = ["出境5", "出境6", "入境10", "入境11", "出境7", "出境8"]
RESCUE_FILL_ALL = True  # if still short after round2, keep pulling until cand exhausted


def _debug(msg: str) -> None:
    if DEBUG_SCHED:
        print(msg)

def validate_input_excel(input_excel_path: str) -> list[dict]:
    """Validate input Excel before scheduling.

    Returns a list of issues:
      [{"sheet": str, "columns": [str, ...], "reason": str}, ...]
    """
    issues: list[dict] = []

    xls_path = os.path.abspath(os.path.expanduser(str(input_excel_path)))
    if not os.path.exists(xls_path):
        return [{"sheet": "檔案", "columns": ["input_excel_path"], "reason": "找不到檔案"}]

    required_sheets = ["排休", "職能", "參數設定"]
    try:
        xls = pd.ExcelFile(xls_path)
        sheet_names = set(xls.sheet_names or [])
    except Exception:
        return [{"sheet": "檔案", "columns": ["Excel"], "reason": "無法讀取Excel"}]

    for sheet in required_sheets:
        if sheet not in sheet_names:
            issues.append({"sheet": sheet, "columns": [], "reason": "缺少工作表"})

    # 排休 sheet checks
    if "排休" in sheet_names:
        try:
            dayoff_raw = pd.read_excel(xls_path, sheet_name="排休")
            cols = [str(c).strip() for c in dayoff_raw.columns]
            sheet_issues: list[str] = []

            def _date_series(df: pd.DataFrame):
                if "日期" not in df.columns:
                    return None
                ds = df["日期"]
                if isinstance(ds, pd.DataFrame):
                    ds = ds.iloc[:, 0]
                return pd.to_numeric(ds, errors="coerce")

            def _detect_data_start_idx(df: pd.DataFrame, emp_cols: list[str]) -> int:
                PART_MARKERS = {"A", "B", "C"}

                def _norm(v) -> str:
                    if pd.isna(v):
                        return ""
                    return str(v).strip()

                def _norm_id(v) -> str:
                    if pd.isna(v):
                        return ""
                    s = str(v).strip()
                    if s == "":
                        return ""
                    try:
                        f = float(s)
                        if f.is_integer():
                            return str(int(f))
                    except Exception:
                        pass
                    return s

                def _looks_like_id(v: str) -> bool:
                    if v == "":
                        return False
                    u = v.upper()
                    if u in PART_MARKERS:
                        return False
                    if any(tok in v for tok in ["休", "請假", "代", "代班"]):
                        return False
                    allowed = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_- ")
                    return all(ch in allowed for ch in u)

                def _score_part_row(row) -> float:
                    vals = [_norm(row.get(emp, "")).upper() for emp in emp_cols]
                    if not vals:
                        return 0.0
                    hit = sum(v in PART_MARKERS for v in vals)
                    return hit / len(vals)

                def _score_id_row(row) -> float:
                    vals = [_norm_id(row.get(emp, "")) for emp in emp_cols]
                    if not vals:
                        return 0.0
                    hit = sum(_looks_like_id(v) for v in vals)
                    return hit / len(vals)

                search_n = 8
                best_part_i, best_part_score = None, -1.0
                best_id_i, best_id_score = None, -1.0

                for i in range(min(search_n, len(df))):
                    row = df.iloc[i]
                    ps = _score_part_row(row)
                    if ps > best_part_score:
                        best_part_score = ps
                        best_part_i = i
                    ids = _score_id_row(row)
                    if ids > best_id_score:
                        best_id_score = ids
                        best_id_i = i

                if best_id_i == best_part_i:
                    best_id_i, best_id_score = None, -1.0
                    for i in range(min(search_n, len(df))):
                        if i == best_part_i:
                            continue
                        ids = _score_id_row(df.iloc[i])
                        if ids > best_id_score:
                            best_id_score = ids
                            best_id_i = i

                default_start = 3 if len(df) > 3 else 0
                if best_part_i is None and best_id_i is None:
                    return default_start
                idxs = [i for i in (best_part_i, best_id_i) if i is not None]
                if not idxs:
                    return default_start
                return max(idxs) + 1

            def _norm_cell(v) -> str:
                if pd.isna(v):
                    return ""
                return str(v).strip()

            if len(cols) < 2:
                sheet_issues.extend(["工作(第1欄)", "日期(第2欄)"])
            else:
                if cols[0] != "工作":
                    sheet_issues.append("工作(第1欄)")
                if cols[1] != "日期":
                    sheet_issues.append("日期(第2欄)")

            if len(cols) <= 2:
                sheet_issues.append("員工欄位(第3欄起)")

            if "日期" in cols:
                dates = _date_series(dayoff_raw)
                if dates is not None and dates.dropna().empty:
                    sheet_issues.append("日期(需有數值)")

            if sheet_issues:
                issues.append(
                    {"sheet": "排休", "columns": sheet_issues, "reason": "欄位不符合模板"}
                )

            # Strict check for dayoff cells: only allow empty / 休 / 代*
            typo_cols: set[str] = set()
            typo_cells: list[dict] = []
            if len(cols) > 2:
                employee_col_info = [(i, cols[i]) for i in range(2, len(cols))]
                data_start_idx = _detect_data_start_idx(dayoff_raw, cols[2:])
                data_block = dayoff_raw.iloc[data_start_idx:].copy() if len(dayoff_raw) > data_start_idx else dayoff_raw.copy()
                dates = _date_series(data_block)
                for col_idx, col_name in employee_col_info:
                    try:
                        s = data_block.iloc[:, col_idx].fillna("").astype(str).str.strip()
                    except Exception:
                        continue
                    if s.eq("").all():
                        continue
                    mask = ~(
                        s.eq("")
                        | s.eq("休")
                        | s.eq("請假")
                        | s.eq("優")
                        | s.str.startswith("代")
                    )
                    if mask.any():
                        typo_cols.add(col_name)
                        if dates is not None:
                            for idx in data_block.index[mask]:
                                try:
                                    d = dates.loc[idx]
                                    if pd.isna(d):
                                        continue
                                    row_pos = data_block.index.get_loc(idx)
                                    if not isinstance(row_pos, int):
                                        continue
                                    val = _norm_cell(data_block.iloc[row_pos, col_idx])
                                    typo_cells.append({"day": int(d), "person": str(col_name), "value": val})
                                except Exception:
                                    continue
            if typo_cols:
                issues.append(
                    {
                        "sheet": "排休",
                        "columns": sorted(typo_cols),
                        "reason": "含不合規字元(僅允許空白/休/請假/優/代開頭)",
                        "cells": typo_cells,
                    }
                )

            # Check 輪休 rows must be empty in employee columns
            if len(cols) > 2:
                employee_col_info = [(i, cols[i]) for i in range(2, len(cols))]
                data_block = dayoff_raw.copy()
                if "工作" in data_block.columns:
                    rest_mask = data_block["工作"].fillna("").astype(str).str.strip() == "輪休"
                    if rest_mask.any():
                        dates = _date_series(data_block)
                        bad_cols: set[str] = set()
                        rest_cells: list[dict] = []
                        for row_idx in data_block.index[rest_mask]:
                            try:
                                row_pos = data_block.index.get_loc(row_idx)
                                if not isinstance(row_pos, int):
                                    continue
                            except Exception:
                                continue
                            for col_idx, col_name in employee_col_info:
                                try:
                                    val = _norm_cell(data_block.iloc[row_pos, col_idx])
                                except Exception:
                                    continue
                                if val == "":
                                    continue
                                bad_cols.add(col_name)
                                if dates is not None:
                                    try:
                                        d = dates.loc[row_idx]
                                        if pd.isna(d):
                                            continue
                                        rest_cells.append({"day": int(d), "person": str(col_name), "value": val})
                                    except Exception:
                                        continue
                        if bad_cols:
                            issues.append(
                                {
                                    "sheet": "排休",
                                    "columns": sorted(bad_cols),
                                    "reason": "輪休天誤植資訊",
                                    "cells": rest_cells,
                                }
                            )
        except Exception:
            issues.append({"sheet": "排休", "columns": [], "reason": "讀取失敗"})

    # 職能 sheet checks
    if "職能" in sheet_names:
        try:
            skill_df = pd.read_excel(xls_path, sheet_name="職能")
            cols = [str(c).strip() for c in skill_df.columns]
            sheet_issues: list[str] = []

            front_cols = cols[: min(7, len(cols))]
            required_shift_cols = ["入境10", "入境11", "出境5", "出境6", "出境7", "出境8"]
            missing_shift_cols = [c for c in required_shift_cols if c not in front_cols]
            sheet_issues.extend(missing_shift_cols)

            name_cols = ["職位", "職能", "技能", "能力"]
            if not any(c in front_cols for c in name_cols):
                sheet_issues.append("職能名稱欄(職位/職能/技能/能力)")

            if len(cols) < 8:
                sheet_issues.append("員工欄位(第8欄起)")

            if sheet_issues:
                issues.append(
                    {"sheet": "職能", "columns": sheet_issues, "reason": "欄位不符合模板"}
                )
        except Exception:
            issues.append({"sheet": "職能", "columns": [], "reason": "讀取失敗"})

    # 參數設定 sheet checks
    if "參數設定" in sheet_names:
        try:
            variables_df = pd.read_excel(xls_path, sheet_name="參數設定")
            cols = [str(c).strip() for c in variables_df.columns]
            sheet_issues: list[str] = []

            required_cols = [
                "第一輪早班？",
                "入境10人數",
                "入境11人數",
                "出境5人數",
                "出境6人數",
                "出境7人數",
                "出境8人數",
            ]
            missing_cols = [c for c in required_cols if c not in cols]
            sheet_issues.extend(missing_cols)

            if "第一輪早班？" in cols and len(variables_df) > 0:
                v = str(variables_df.iloc[0].get("第一輪早班？", "")).strip()
                if v not in ("A", "B"):
                    sheet_issues.append("第一輪早班？(值需A/B)")

            for col in [
                "入境10人數",
                "入境11人數",
                "出境5人數",
                "出境6人數",
                "出境7人數",
                "出境8人數",
            ]:
                if col in cols and len(variables_df) > 0:
                    v = variables_df.iloc[0].get(col, None)
                    num = pd.to_numeric(v, errors="coerce")
                    if pd.isna(num):
                        sheet_issues.append(f"{col}(需為數字)")

            if sheet_issues:
                issues.append(
                    {"sheet": "參數設定", "columns": sheet_issues, "reason": "欄位不符合模板"}
                )
        except Exception:
            issues.append({"sheet": "參數設定", "columns": [], "reason": "讀取失敗"})

    return issues

def reset_schedule_state(daily_list: list[dict], people_dict: dict) -> None:
    """Hard reset all mutable scheduling state.

    Use this before re-running *any* scheduling loops in the same Python session,
    especially when you run code blocks in batches in VSCode.

    Resets:
    - each day: _cand, 特殊職務, and each shift list under 班段
    - each employee: 拉班次數, 公平性分數, 班段次數 (set all existing keys back to 0)
    """

    # 1) Reset per-day state
    for dd in daily_list:
        if not isinstance(dd, dict):
            continue

        # Remove leftover candidate list from previous runs
        if "_cand" in dd:
            dd.pop("_cand", None)

        # Remove special duty assignments
        if "特殊職務" in dd:
            dd.pop("特殊職務", None)

        # Clear shift assignments but keep the 班段 keys that exist for that day
        bd = dd.get("班段", {})
        if isinstance(bd, dict):
            for sh in list(bd.keys()):
                # Only clear list-type shifts; leave {} for 輪休 days as-is
                if isinstance(bd.get(sh), list):
                    bd[sh] = []

    # 2) Reset per-employee counters
    for emp, info in people_dict.items():
        if not isinstance(info, dict):
            continue
        info["拉班次數"] = 0

        fairness = info.get("公平性分數", {})
        if isinstance(fairness, dict):
            # Keep keys but reset values
            for k in list(fairness.keys()):
                fairness[k] = 0
            info["公平性分數"] = fairness

        shift_counts = info.get("班段次數", {})
        if isinstance(shift_counts, dict):
            for k in list(shift_counts.keys()):
                shift_counts[k] = 0
            info["班段次數"] = shift_counts


def debug_state_snapshot(daily_list: list[dict], people_dict: dict, days_n: int = 5) -> None:
    """Print a quick snapshot to detect state leakage."""
    if not DEBUG_SCHED:
        return

    # Cand snapshot
    cands = []
    for dd in daily_list:
        if not isinstance(dd, dict):
            continue
        if dd.get("班段") and "_cand" in dd:
            cands.append((int(dd.get("日期", 0) or 0), len(dd.get("_cand", []) or [])))
    if cands:
        mn = min(cands, key=lambda x: x[1])
        mx = max(cands, key=lambda x: x[1])
        _debug(f"[SNAPSHOT] _cand days={len(cands)} min={mn} max={mx}")

    # Pull count snapshot
    pulls = []
    for emp, info in people_dict.items():
        try:
            pulls.append((emp, int(info.get("拉班次數", 0) or 0)))
        except Exception:
            pulls.append((emp, 0))
    if pulls:
        pulls_sorted = sorted(pulls, key=lambda x: x[1], reverse=True)
        top = pulls_sorted[:10]
        _debug("[SNAPSHOT] 拉班次數 Top10: " + ", ".join([f"{e}:{n}" for e, n in top]))

    # Assigned count snapshot (first few days)
    shown = 0
    for dd in daily_list:
        if not isinstance(dd, dict):
            continue
        if not dd.get("班段"):
            continue
        d = int(dd.get("日期", 0) or 0)
        total_assigned = 0
        for sh, recs in (dd.get("班段", {}) or {}).items():
            if isinstance(recs, list):
                total_assigned += len(recs)
        _debug(f"[SNAPSHOT] day {d}: total_assigned={total_assigned}")
        shown += 1
        if shown >= days_n:
            break


#
# =========================
# Web-friendly runtime state
# (populated by run_scheduler; do NOT run at import-time)
# =========================

dayoff_raw: pd.DataFrame | None = None
skill: pd.DataFrame | None = None
variables: pd.DataFrame | None = None

dayoff: pd.DataFrame | None = None
people_dict: dict = {}
employee_cols = []
daily_list: list[dict] = []
shift_demands: dict[str, int] = {}

# Default limit used by scheduling/output helpers
DAYS_LIMIT = 28

# Optional overrides for pipeline/scorers
TEAM_PIPELINE_OVERRIDE: dict | None = None
TEAM_SCORERS_NONPULL_OVERRIDE: dict | None = None


def run_scheduler(
    input_excel_path: str,
    output_excel_path: str | None = None,
    *,
    days_limit: int | None = None,
    include_external: bool = False,
    search_best_roster: bool = True,
    search_patience: int = 10,
    require_all_pulls_nonzero: bool = True,
    reset_before_sched: bool = True,
    smart_team_pick: bool = True,
    debug: bool = False,
    random_seed: int | None = None,
    progress_callback=None,
    priority_mode: str = "team1",
    custom_order: str = "fairness,shift_count",
    rescue_fill: bool = True,
    score_order: str = "fairness,shift,pull",
) -> dict:
    """Run scheduler from an uploaded Excel and export an output Excel.

    Web usage:
      result = run_scheduler(uploaded_path)
      # result['output_path'] is the file to send back to the user.

    Returns:
      {
        'output_path': str,
        'tries': int,
        'best_std': float,        # combined score: pull_std + fair_std + shift_std (negated)
        'best_pull_std': float,   # pull-count std (A/B)
        'best_fair_std': float,   # fairness-sum std (A/B)
        'best_shift_std': float,  # shift-count std sum (A/B)
        'best_score_100': float,  # normalized 0-100 score (higher is better)
        'used_search': bool,
      }
    """

    import tempfile

    global dayoff_raw, skill, variables, dayoff
    global people_dict, employee_cols, daily_list, shift_demands
    global DAYS_LIMIT, DEBUG_SCHED, RESET_BEFORE_SCHED, SMART_TEAM_PICK
    global SEARCH_BEST_ROSTER, SEARCH_MAX_TRIES, SEARCH_MIN_TRIES, SEARCH_PATIENCE
    global TEAM_PIPELINE_OVERRIDE, TEAM_SCORERS_NONPULL_OVERRIDE
    global RESCUE_FILL_ALL

    # Apply run-time switches
    DEBUG_SCHED = bool(debug)
    RESET_BEFORE_SCHED = bool(reset_before_sched)
    SMART_TEAM_PICK = False

    SEARCH_BEST_ROSTER = bool(search_best_roster)
    SEARCH_MIN_TRIES = 100
    SEARCH_MAX_TRIES = 100
    SEARCH_PATIENCE = int(search_patience)
    RESCUE_FILL_ALL = bool(rescue_fill)

    # Configure scorer priority based on mode
    priority_mode = str(priority_mode or "").strip().lower()
    order = [s.strip() for s in str(custom_order or "").split(",") if s.strip()]
    active_set = {k for k in order if k in ("fairness", "shift_count")}
    if not active_set:
        active_set = {"fairness", "shift_count"}

    # Global score priority order (weights 3,2,1)
    score_order_list = [s.strip() for s in str(score_order or "").split(",") if s.strip()]
    if len(score_order_list) != 3:
        score_order_list = ["fairness", "shift", "pull"]
    if priority_mode == "team1":
        score_order_list = ["fairness", "pull", "shift"]

    def _make_nonpull_scorers(order_list: list[str]):
        mapping = {
            "fairness": "_sc_fairness",
            "shift_count": "_sc_shift_count",
        }
        return [mapping.get(k, "_sc_fairness") for k in order_list]

    if priority_mode == "custom":
        def _nonpull_order_from_score(score_list: list[str]) -> list[str]:
            out: list[str] = []
            for k in score_list:
                if k == "fairness" and "fairness" in active_set:
                    out.append("fairness")
                elif k == "shift" and "shift_count" in active_set:
                    out.append("shift_count")
            return out

        custom_nonpull = _nonpull_order_from_score(score_order_list)
        if not custom_nonpull:
            custom_nonpull = ["fairness"] if "fairness" in active_set else ["shift_count"]

        TEAM_SCORERS_NONPULL_OVERRIDE = {
            "A": custom_nonpull,
            "B": custom_nonpull,
            "C": custom_nonpull,
        }
    elif priority_mode == "team3":
        TEAM_SCORERS_NONPULL_OVERRIDE = {
            "A": ["shift_count", "fairness"],
            "B": ["shift_count", "fairness"],
            "C": ["shift_count", "fairness"],
        }
    else:
        # team1 / team2 default: fairness before shift count
        TEAM_SCORERS_NONPULL_OVERRIDE = {
            "A": ["fairness", "shift_count"],
            "B": ["fairness", "shift_count"],
            "C": ["fairness", "shift_count"],
        }

    if days_limit is not None:
        DAYS_LIMIT = int(days_limit)
        if DAYS_LIMIT < 1:
            raise ValueError("days_limit must be >= 1")

    if random_seed is not None:
        try:
            random.seed(int(random_seed))
        except Exception:
            pass

    # Resolve paths
    xls_path = os.path.abspath(os.path.expanduser(str(input_excel_path)))
    if not os.path.exists(xls_path):
        raise FileNotFoundError(f"Input Excel not found: {xls_path}")

    if output_excel_path is None or str(output_excel_path).strip() == "":
        tmpdir = tempfile.mkdtemp(prefix="airport_roster_")
        out_path = os.path.join(tmpdir, "roster_output.xlsx")
    else:
        out_path = os.path.abspath(os.path.expanduser(str(output_excel_path)))
        os.makedirs(os.path.dirname(out_path), exist_ok=True)

    # -------------------------
    # Load Excel (no Desktop hard-coding)
    # -------------------------
    dayoff_raw = pd.read_excel(xls_path, sheet_name="排休")
    skill = pd.read_excel(xls_path, sheet_name="職能")
    variables = pd.read_excel(xls_path, sheet_name="參數設定")

    # Build people_dict shell
    people_dict = {}
    employee_cols = dayoff_raw.columns[2:]

    PART_MARKERS = {"A", "B", "C"}

    def _norm(v) -> str:
        if pd.isna(v):
            return ""
        return str(v).strip()

    def _norm_id(v) -> str:
        if pd.isna(v):
            return ""
        s = str(v).strip()
        if s == "":
            return ""
        try:
            f = float(s)
            if f.is_integer():
                return str(int(f))
        except Exception:
            pass
        return s

    def _score_part_row(row) -> float:
        vals = [_norm(row.get(emp, "")).upper() for emp in employee_cols]
        if not vals:
            return 0.0
        hit = sum(v in PART_MARKERS for v in vals)
        return hit / len(vals)

    def _looks_like_id(v: str) -> bool:
        if v == "":
            return False
        u = v.upper()
        if u in PART_MARKERS:
            return False
        if any(tok in v for tok in ["休", "請假", "代", "代班"]):
            return False
        allowed = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_- ")
        return all(ch in allowed for ch in u)

    def _score_id_row(row) -> float:
        vals = [_norm(row.get(emp, "")) for emp in employee_cols]
        if not vals:
            return 0.0
        hit = sum(_looks_like_id(v) for v in vals)
        return hit / len(vals)

    # Auto-detect part row and id row
    search_n = 8
    best_part_i, best_part_score = None, -1.0
    best_id_i, best_id_score = None, -1.0

    for i in range(min(search_n, len(dayoff_raw))):
        row = dayoff_raw.iloc[i]
        ps = _score_part_row(row)
        if ps > best_part_score:
            best_part_score = ps
            best_part_i = i

        ids = _score_id_row(row)
        if ids > best_id_score:
            best_id_score = ids
            best_id_i = i

    if best_id_i == best_part_i:
        best_id_i, best_id_score = None, -1.0
        for i in range(min(search_n, len(dayoff_raw))):
            if i == best_part_i:
                continue
            ids = _score_id_row(dayoff_raw.iloc[i])
            if ids > best_id_score:
                best_id_score = ids
                best_id_i = i

    data_start_idx = 3
    if best_part_i is not None or best_id_i is not None:
        idxs = [i for i in (best_part_i, best_id_i) if i is not None]
        if idxs:
            data_start_idx = max(idxs) + 1

    # -------------------------
    # Parse dayoff (after detecting header rows)
    # -------------------------
    dayoff = dayoff_raw.iloc[data_start_idx:].copy()
    dayoff["日期"] = pd.to_numeric(dayoff["日期"], errors="coerce")
    dayoff = dayoff.dropna(subset=["日期"]).copy()
    dayoff["日期"] = dayoff["日期"].astype(int)

    part_row = dayoff_raw.iloc[best_part_i] if best_part_i is not None else dayoff_raw.iloc[0]
    id_row = dayoff_raw.iloc[best_id_i] if best_id_i is not None else dayoff_raw.iloc[0]

    emp_to_part: dict[str, str] = {}
    emp_to_id: dict[str, str] = {}

    for emp in employee_cols:
        emp_to_part[emp] = _norm(part_row.get(emp, "")).upper()
        emp_to_id[emp] = _norm_id(id_row.get(emp, ""))

    OFF_MARKERS = {"休", "請假", "優"}
    SUB_MARKERS = {"代", "代班"}

    for emp in employee_cols:
        tmp = dayoff[["日期", "工作", emp]].copy()
        tmp[emp] = tmp[emp].fillna("").astype(str).str.strip()

        off_days: list[int] = []
        sub_pairs: list[tuple[int, str]] = []

        for _, r in tmp.iterrows():
            day_i = int(r["日期"])
            val = r[emp]

            if val == "":
                continue

            if val in OFF_MARKERS:
                off_days.append(day_i)
                continue

            if (val in SUB_MARKERS) or val.startswith("代"):
                helper = str(val)
                if helper.startswith("代"):
                    helper = helper[1:]
                helper = helper.strip().strip("()（） ")
                sub_pairs.append((day_i, helper))
                continue

        def _dedupe_keep_order(nums: list[int]) -> list[int]:
            seen = set()
            out = []
            for x in nums:
                if x not in seen:
                    seen.add(x)
                    out.append(x)
            return out

        people_dict[emp] = {
            "ID": emp_to_id.get(emp, ""),
            "分組": emp_to_part.get(emp, ""),
            "休假": _dedupe_keep_order(off_days),
            "代班": {},
            "代班日期": [],
            "代班人員": [],
            "職能": {},
            "公平性分數": {},
            "拉班次數": 0,
            "班段次數": {sh: 0 for sh in SHIFT_ORDER},
        }

        sub_map: dict[int, str] = {}
        for day_i, helper_i in sub_pairs:
            if day_i not in sub_map:
                sub_map[day_i] = helper_i
            else:
                if (sub_map[day_i] == "") and (helper_i != ""):
                    sub_map[day_i] = helper_i

        people_dict[emp]["代班"] = sub_map
        people_dict[emp]["代班日期"] = list(sub_map.keys())
        people_dict[emp]["代班人員"] = [sub_map[k] for k in sub_map.keys()]

    # -------------------------
    # Skill matrix mapping (same behavior, but raise instead of print)
    # -------------------------
    id_to_emp: dict[str, str] = {}
    for emp in employee_cols:
        eid = _norm_id(people_dict[emp].get("ID", ""))
        if eid:
            id_to_emp[eid] = emp
            people_dict[emp]["ID"] = eid

    if len(id_to_emp) == 0:
        raise ValueError("[ERROR] id_to_emp is empty. Check dayoff ID extraction.")

    front_cols = list(skill.columns[: min(7, len(skill.columns))])

    def _is_texty(x: str) -> bool:
        if x == "" or x.lower() == "nan":
            return False
        if re.fullmatch(r"\d+(\.\d+)?", x):
            return False
        if len(x) <= 1 and not any(ch.isalpha() or "\u4e00" <= ch <= "\u9fff" for ch in x):
            return False
        return True

    best_col = front_cols[0] if front_cols else skill.columns[0]
    best_score = -1

    for c in front_cols:
        s = skill[c].fillna("").astype(str).str.strip()
        score = sum(_is_texty(v) for v in s.tolist())
        if score > best_score:
            best_score = score
            best_col = c

    skill_name_col = best_col
    skill_emp_cols = list(skill.columns[7:])

    name_series = skill[skill_name_col].fillna("").astype(str).str.strip()
    is_blank = name_series.eq("") | name_series.str.lower().eq("nan")
    is_numeric = name_series.str.fullmatch(r"\d+(\.\d+)?").fillna(False)
    valid_skill_rows = skill[~is_blank & ~is_numeric]

    search_skill_id_n = 12
    best_skill_id_i, best_skill_id_score = None, -1.0
    known_ids = set(id_to_emp.keys())

    for i in range(min(search_skill_id_n, len(skill))):
        row = skill.iloc[i]
        vals = [_norm_id(row.get(c, "")) for c in skill_emp_cols]
        if not vals:
            continue
        hit = sum(v in known_ids for v in vals)
        score = hit / len(vals)
        if score > best_skill_id_score:
            best_skill_id_score = score
            best_skill_id_i = i

    skill_id_row = skill.iloc[best_skill_id_i] if best_skill_id_i is not None else skill.iloc[0]

    unmatched_skill_ids: list[str] = []

    for col in skill_emp_cols:
        sid = _norm_id(skill_id_row.get(col, ""))
        if not sid:
            continue

        emp = id_to_emp.get(sid)
        if not emp:
            unmatched_skill_ids.append(sid)
            continue

        emp_skills: dict[str, bool] = {}
        for _, row in valid_skill_rows.iterrows():
            sk_name = _norm(row.get(skill_name_col, ""))
            if not sk_name:
                continue

            v = row.get(col, None)
            if pd.isna(v):
                continue

            is_one = False
            try:
                is_one = int(v) == 1
            except Exception:
                is_one = str(v).strip() == "1"

            if not is_one:
                continue

            emp_skills[sk_name] = True

        people_dict[emp]["職能"] = emp_skills

        fairness = people_dict[emp].get("公平性分數", {})
        if not isinstance(fairness, dict):
            fairness = {}
        for sk in emp_skills.keys():
            fairness.setdefault(sk, 0)
        people_dict[emp]["公平性分數"] = fairness

    # -------------------------
    # Level 1 skeleton (build daily_list)
    # -------------------------
    base = dayoff_raw.iloc[:, :2].copy()
    base.columns = ["工作", "日期"]

    base["日期"] = pd.to_numeric(base["日期"], errors="coerce")
    mask_na = base["日期"].isna()
    if mask_na.any():
        raw_str = base.loc[mask_na, "日期"].astype(str).str.strip()
        extracted = raw_str.str.extract(r"(\d+)")[0]
        base.loc[mask_na, "日期"] = pd.to_numeric(extracted, errors="coerce")

    base = base.dropna(subset=["日期"]).copy()
    base["日期"] = base["日期"].astype(int)
    base["工作"] = base["工作"].fillna("").astype(str).str.strip()

    all_days: list[int] = sorted(base["日期"].dropna().unique().tolist())

    def _infer_day_type(works: list[str]) -> str:
        works_norm = [str(w).strip() for w in works if str(w).strip() != ""]
        has_in = any(("入境" in w) or w.startswith("入") for w in works_norm)
        has_out = any(("出境" in w) or w.startswith("出") for w in works_norm)
        has_rest = any("輪休" in w for w in works_norm)

        if has_in and not has_out:
            return "入境"
        if has_out and not has_in:
            return "出境"
        if has_in and has_out:
            return "混合"
        if has_rest:
            return "輪休"
        return "未知"

    # Use your variables sheet to get first-team + demands
    first_team, shift_demands = get_shift_demands(variables)
        # Validate demands to avoid KeyErrors later
    if str(first_team).strip() not in ("A", "B"):
        raise ValueError(f"[ERROR] 第一輪早班？ must be 'A' or 'B', got: {first_team!r}")

    missing_demands = [sh for sh in SHIFT_ORDER if sh not in (shift_demands or {})]
    if missing_demands:
        raise ValueError(f"[ERROR] 參數設定缺少班段需求欄位: {missing_demands}")

    TeamChoices = ["A", "B"]
    First_part_starting_team = str(first_team)
    part = 0
    if First_part_starting_team in TeamChoices:
        TeamChoices.remove(First_part_starting_team)
    AnotherChoice = str(TeamChoices[0]) if TeamChoices else ""

    # Determine scheduling window:
    # start from first 入境 day; end at last 出境 day in the last part.
    day_type_by_day: dict[int, str] = {}
    part_by_day: dict[int, int] = {}
    part_idx = 0
    for d in all_days:
        works = base.loc[base["日期"] == int(d), "工作"].tolist()
        day_type = _infer_day_type(works)
        day_type_by_day[int(d)] = day_type
        part_by_day[int(d)] = part_idx
        if day_type == "輪休":
            part_idx += 1

    start_day = None
    for d in all_days:
        if day_type_by_day.get(int(d)) == "入境":
            start_day = int(d)
            break
    if start_day is None and all_days:
        start_day = int(all_days[0])

    last_part = max(part_by_day.values()) if part_by_day else 0
    end_day = None
    for d in reversed(all_days):
        if part_by_day.get(int(d)) == last_part and day_type_by_day.get(int(d)) == "出境":
            end_day = int(d)
            break
    if end_day is None and all_days:
        for d in reversed(all_days):
            if part_by_day.get(int(d)) == last_part:
                end_day = int(d)
                break

    schedule_days = [d for d in all_days if start_day is not None and end_day is not None and start_day <= int(d) <= end_day]
    if not schedule_days:
        schedule_days = all_days[:]

    days: dict[int, dict] = {}
    for d in schedule_days:
        works = base.loc[base["日期"] == int(d), "工作"].tolist()
        day_type = _infer_day_type(works)

        day_dict = {"日期": int(d)}

        if day_type == "入境":
            day_dict["班段"] = {"入境10": [], "入境11": []}
            day_dict["早班"] = First_part_starting_team if part % 2 == 0 else AnotherChoice
        elif day_type == "出境":
            day_dict["班段"] = {"出境5": [], "出境6": [], "出境7": [], "出境8": []}
            day_dict["早班"] = AnotherChoice if part % 2 == 0 else First_part_starting_team
        elif day_type == "輪休":
            part += 1
            day_dict["班段"] = {}
            day_dict["早班"] = {}
        else:
            # Unknown day type: still create a shell to avoid crashing
            day_dict["班段"] = {}
            day_dict["早班"] = {}

        days[int(d)] = day_dict

    daily_list = [days[d] for d in schedule_days]

    if days_limit is None:
        DAYS_LIMIT = len(daily_list)

    # -------------------------
    # Run scheduling (search-best or single)
    # -------------------------
    used_search = bool(search_best_roster)
    # Score: higher is better (we use negative std sum)
    best_score = float("-inf")
    best_pull_std = float("inf")
    best_fair_std = float("inf")
    best_shift_std = float("inf")
    best_daily = None
    best_people = None
    best_is_bad = False

    best_good_score = float("-inf")
    best_good_pull_std = float("inf")
    best_good_fair_std = float("inf")
    best_good_shift_std = float("inf")
    best_good_daily = None
    best_good_people = None

    best_bad_score = float("-inf")
    best_bad_pull_std = float("inf")
    best_bad_fair_std = float("inf")
    best_bad_shift_std = float("inf")
    best_bad_daily = None
    best_bad_people = None
    no_improve = 0
    total_tries = 0
    # If many attempts are skipped (RULE violations / pulls-nonzero constraint),
    # the loop can take a long time while making no progress. Guard with a skip streak.
    skip_streak = 0

    if search_best_roster:
        for t in range(1, SEARCH_MIN_TRIES + 1):
            total_tries = t
            if callable(progress_callback):
                try:
                    progress_callback(t, SEARCH_MIN_TRIES)
                except Exception:
                    pass
            try:
                _schedule_once()
            except ValueError as e:
                if str(e).startswith("[RULE]"):
                    skip_streak += 1
                    continue
                skip_streak += 1
                continue
            except Exception:
                skip_streak += 1
                continue

            # Post-fix module before scoring
            _repair_in11_shortage(daily_list, people_dict, employee_cols)

            if require_all_pulls_nonzero and not _all_pulls_nonzero_ab(people_dict):
                skip_streak += 1
                continue

            # We have a valid roster attempt; reset skip streak.
            skip_streak = 0

            pull_std = _pull_std_ab(people_dict)
            fair_std = _fairness_sum_std_ab(people_dict)
            shift_std = _shift_count_std_ab(people_dict)
            std_map = {"pull": pull_std, "fairness": fair_std, "shift": shift_std}
            weights = [3, 2, 1]
            weighted_sum = 0.0
            for w, k in zip(weights, score_order_list):
                weighted_sum += float(std_map.get(k, 0.0)) * w
            score = -(weighted_sum)

            has_empty = _violates_no_empty_on_workday(daily_list, employee_cols, people_dict)
            if not has_empty:
                if score > best_good_score + 1e-12:
                    best_good_score = score
                    best_good_pull_std = pull_std
                    best_good_fair_std = fair_std
                    best_good_shift_std = shift_std
                    best_good_daily = copy.deepcopy(daily_list[:DAYS_LIMIT])
                    best_good_people = copy.deepcopy(people_dict)
                    no_improve = 0
                else:
                    no_improve += 1
            else:
                if score > best_bad_score + 1e-12:
                    best_bad_score = score
                    best_bad_pull_std = pull_std
                    best_bad_fair_std = fair_std
                    best_bad_shift_std = shift_std
                    best_bad_daily = copy.deepcopy(daily_list[:DAYS_LIMIT])
                    best_bad_people = copy.deepcopy(people_dict)
                no_improve += 1

        if best_good_daily is not None and best_good_people is not None:
            best_score = best_good_score
            best_pull_std = best_good_pull_std
            best_fair_std = best_good_fair_std
            best_shift_std = best_good_shift_std
            best_daily = best_good_daily
            best_people = best_good_people
            best_is_bad = False
        elif best_bad_daily is not None and best_bad_people is not None:
            best_score = best_bad_score
            best_pull_std = best_bad_pull_std
            best_fair_std = best_bad_fair_std
            best_shift_std = best_bad_shift_std
            best_daily = best_bad_daily
            best_people = best_bad_people
            best_is_bad = True

        if best_daily is None or best_people is None:
            raise ValueError(
                f"[FINAL] No valid roster found within {total_tries} tries (min={SEARCH_MIN_TRIES})."
            )

        daily_list[:DAYS_LIMIT] = best_daily
        people_dict.clear()
        people_dict.update(best_people)

    else:
        total_tries = 1
        if callable(progress_callback):
            try:
                progress_callback(1, 1)
            except Exception:
                pass
        _schedule_once()
        _repair_in11_shortage(daily_list, people_dict, employee_cols)
        if require_all_pulls_nonzero and not _all_pulls_nonzero_ab(people_dict):
            raise ValueError("[FINAL] Single run violated constraint: some A/B 拉班次數 is 0")
        best_pull_std = _pull_std_ab(people_dict)
        best_fair_std = _fairness_sum_std_ab(people_dict)
        best_shift_std = _shift_count_std_ab(people_dict)
        std_map = {"pull": best_pull_std, "fairness": best_fair_std, "shift": best_shift_std}
        weights = [3, 2, 1]
        weighted_sum = 0.0
        for w, k in zip(weights, score_order_list):
            weighted_sum += float(std_map.get(k, 0.0)) * w
        best_score = -(weighted_sum)

    # -------------------------
    # Build output df + export to Excel
    # -------------------------
    roster_df = build_roster_table(daily_list[:DAYS_LIMIT], employee_cols, people_dict, include_external=include_external)

    def _id_key(emp: str):
        v = str(people_dict.get(emp, {}).get("ID", "") or "").strip()
        try:
            return (0, int(v))
        except Exception:
            return (1, v)

    leaders = [e for e in employee_cols if people_dict.get(e, {}).get("分組", "") == "C"]
    A_team = [e for e in employee_cols if people_dict.get(e, {}).get("分組", "") == "A"]
    B_team = [e for e in employee_cols if people_dict.get(e, {}).get("分組", "") == "B"]

    leaders = sorted(leaders, key=_id_key)
    A_team = sorted(A_team, key=_id_key)
    B_team = sorted(B_team, key=_id_key)

    ordered_cols = leaders + A_team + B_team
    rest = [e for e in roster_df.columns if e not in ordered_cols]
    ordered_cols += rest

    roster_df = roster_df[ordered_cols]

    rest_days = set()
    for dd in daily_list[:DAYS_LIMIT]:
        if not dd.get("班段"):
            try:
                rest_days.add(int(dd.get("日期", 0) or 0))
            except Exception:
                pass

    # Precompute pulled cells for coloring (use 原員工)
    pulled_cells: set[tuple[int, str]] = set()
    for dd in daily_list[:DAYS_LIMIT]:
        if not isinstance(dd, dict):
            continue
        d = int(dd.get("日期", 0) or 0)
        for _, recs in (dd.get("班段", {}) or {}).items():
            if not isinstance(recs, list):
                continue
            for rec in recs:
                if not isinstance(rec, dict):
                    continue
                if not bool(rec.get("拉班", False)):
                    continue
                name = str(rec.get("原員工", rec.get("人員", "")) or "").strip()
                if name:
                    pulled_cells.add((d, name))

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        roster_df.to_excel(writer, sheet_name="Roster")
        ws = writer.sheets["Roster"]

        fill_blue = PatternFill("solid", fgColor="9DC3E6")
        fill_yellow = PatternFill("solid", fgColor="FFF2CC")
        fill_green = PatternFill("solid", fgColor="C6E0B4")
        fill_gray = PatternFill("solid", fgColor="E7E6E6")
        fill_pink = PatternFill("solid", fgColor="F8CBAD")
        fill_red_empty = PatternFill("solid", fgColor="F4CCCC")

        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_row = 1

        for j, emp in enumerate(ordered_cols, start=2):
            cell = ws.cell(row=header_row, column=j)
            cell.font = header_font
            cell.alignment = center

            grp = people_dict.get(emp, {}).get("分組", "")
            if grp == "B":
                cell.fill = fill_green
            elif grp == "A":
                cell.fill = fill_yellow
            else:
                cell.fill = fill_blue

        ws.cell(row=header_row, column=1).font = header_font
        ws.cell(row=header_row, column=1).alignment = center

        for i, day in enumerate(roster_df.index.tolist(), start=2):
            max_col = 1 + len(ordered_cols)
            is_rest = int(day) in rest_days

            for col in range(1, max_col + 1):
                c = ws.cell(row=i, column=col)
                c.alignment = center
                if is_rest:
                    c.fill = fill_gray

            if is_rest:
                continue

            for j in range(2, max_col + 1):
                c = ws.cell(row=i, column=j)
                name = str(roster_df.columns[j - 2])
                val = str(c.value or "")
                if (int(day), name) in pulled_cells:
                    c.fill = fill_pink
                if best_is_bad and val == "":
                    c.fill = fill_red_empty

        ws.row_dimensions[1].height = 22

        # Set all columns to the same width (max width as baseline)
        max_width = 0
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            w = ws.column_dimensions[letter].width or 10
            if w > max_width:
                max_width = w
        if max_width <= 0:
            max_width = 12
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = max_width

    # Normalize score to 0-100 (higher is better)
    std_map = {"pull": best_pull_std, "fairness": best_fair_std, "shift": best_shift_std}
    weights = [3, 2, 1]
    raw_total = 0.0
    for w, k in zip(weights, score_order_list):
        raw_total += float(std_map.get(k, 0.0)) * w
    norm = raw_total / (raw_total + 1.0) if raw_total >= 0 else 1.0
    best_score_100 = max(0.0, min(100.0, (1.0 - norm) * 100.0))

    # Build chart data for A/B shift counts
    chart_data: dict = {"shift": {}, "skill": {}, "pull": {}}
    shift_order = ["入境10", "入境11", "出境5", "出境6", "出境7", "出境8"]
    for sh in shift_order:
        a_list = []
        b_list = []
        for emp in employee_cols:
            info = people_dict.get(emp, {})
            grp = str(info.get("分組", "") or "")
            counts = info.get("班段次數", {})
            if not isinstance(counts, dict):
                continue
            c = int(counts.get(sh, 0) or 0)
            if grp == "A":
                a_list.append((emp, c))
            elif grp == "B":
                b_list.append((emp, c))
        a_list.sort(key=lambda x: x[1], reverse=True)
        b_list.sort(key=lambda x: x[1], reverse=True)
        chart_data["shift"][sh] = {"A": a_list, "B": b_list}

    # Skill/role charts (only count employees who have that skill)
    skill_names = set()
    for emp in employee_cols:
        info = people_dict.get(emp, {})
        skills = info.get("職能", {}) or {}
        if isinstance(skills, dict):
            for sk in skills.keys():
                skill_names.add(str(sk))
    skip_skills = {"分隊長", "代理分隊長"}
    for sk in sorted(skill_names):
        if sk in skip_skills:
            continue
        a_list = []
        b_list = []
        for emp in employee_cols:
            info = people_dict.get(emp, {})
            grp = str(info.get("分組", "") or "")
            skills = info.get("職能", {}) or {}
            if not isinstance(skills, dict) or sk not in skills:
                continue
            fairness = info.get("公平性分數", {}) or {}
            if not isinstance(fairness, dict):
                continue
            c = int(fairness.get(sk, 0) or 0)
            if grp == "A":
                a_list.append((emp, c))
            elif grp == "B":
                b_list.append((emp, c))
        a_list.sort(key=lambda x: x[1], reverse=True)
        b_list.sort(key=lambda x: x[1], reverse=True)
        chart_data["skill"][sk] = {"A": a_list, "B": b_list}

    # Pull count chart (A/B)
    a_list = []
    b_list = []
    for emp in employee_cols:
        info = people_dict.get(emp, {})
        grp = str(info.get("分組", "") or "")
        c = int(info.get("拉班次數", 0) or 0)
        if grp == "A":
            a_list.append((emp, c))
        elif grp == "B":
            b_list.append((emp, c))
    a_list.sort(key=lambda x: x[1], reverse=True)
    b_list.sort(key=lambda x: x[1], reverse=True)
    chart_data["pull"]["拉班次數"] = {"A": a_list, "B": b_list}

    return {
        "output_path": out_path,
        "tries": int(total_tries),
        "best_std": float(best_score),
        "best_pull_std": float(best_pull_std),
        "best_fair_std": float(best_fair_std),
        "best_shift_std": float(best_shift_std),
        "best_score_100": float(best_score_100),
        "chart_data": chart_data,
        "best_is_bad": bool(best_is_bad),
        "used_search": bool(used_search),
    }

def get_candidates_for_day(day_dict: dict, employee_cols, people_dict) -> list[str]:
    """
    給定某一天的 day_dict，回傳當天可上班的候選人（排除休假者）
    Given a day_dict, return available employees (exclude those who are off).
    """
    # 1) 讀取日期（強制轉 int，避免 '2' / 2 型別不一致）
    d = int(day_dict["日期"])

    candidates: list[str] = []
    if len(day_dict["班段"]) != 0:
        # 2) 掃每位員工
        for emp in employee_cols:
            emp_info = people_dict.get(emp, {})
            off_days = emp_info.get("休假", [])
                
            # 3) 如果沒休假 -> 進候選
            if d not in off_days:
                candidates.append(emp)

    return candidates


def _is_sub_day(emp: str, d: int, people_dict: dict) -> bool:
    """True if emp has 代班 on day d (meaning they are not actually working that day)."""
    info = people_dict.get(emp, {}) if isinstance(people_dict, dict) else {}

    # Preferred structure: 代班 = {date: helper_name}
    sub_map = info.get("代班", {})
    if isinstance(sub_map, dict):
        try:
            if int(d) in sub_map:
                return True
        except Exception:
            pass

    # Backward-compatible structure
    sub_days = info.get("代班日期", []) or []
    for x in sub_days:
        try:
            if int(x) == int(d):
                return True
        except Exception:
            continue

    return False

def _next_day_dict(d: int) -> dict | None:
    """Get next calendar day's dict from global daily_list (may be None)."""
    try:
        nd = int(d) + 1
    except Exception:
        return None
    for x in (daily_list[:DAYS_LIMIT] or []):
        try:
            if int(x.get("日期", 0) or 0) == nd:
                return x
        except Exception:
            continue
    return None


def _prev_day_dict(d: int) -> dict | None:
    """Get previous calendar day's dict from global daily_list (may be None)."""
    try:
        pd = int(d) - 1
    except Exception:
        return None
    for x in (daily_list[:DAYS_LIMIT] or []):
        try:
            if int(x.get("日期", 0) or 0) == pd:
                return x
        except Exception:
            continue
    return None


def _has_out_early(emp: str, day_dict: dict | None) -> bool:
    """True if emp is assigned to 出境5/出境6 in the given day_dict."""
    if not day_dict or not isinstance(day_dict, dict):
        return False
    bd = day_dict.get("班段", {}) or {}
    for sh in ("出境5", "出境6"):
        recs = bd.get(sh, []) or []
        if not isinstance(recs, list):
            continue
        for r in recs:
            if not isinstance(r, dict):
                continue
            who = str(r.get("原員工", r.get("人員", "")) or "").strip()
            if who == emp:
                return True
    return False


def _violates_in11_out_early(emp: str, d: int, people_dict: dict) -> bool:
    """入境11 不能接隔天出境早班(出境5/6)，除非其中一天是代班日。"""
    if _is_sub_day(emp, d, people_dict) or _is_sub_day(emp, d + 1, people_dict):
        return False
    nd = _next_day_dict(d)
    return _has_out_early(emp, nd)

def _validate_hard_rules_in11_out_early(daily_list: list[dict], people_dict: dict) -> None:
    """Hard rule: 入境11 cannot be followed by next-day 出境早班(出境5/6) for the same person,
    unless either day is a substitute day (代班).
    """
    for dd in (daily_list[:DAYS_LIMIT] or []):
        if not isinstance(dd, dict):
            continue
        d = int(dd.get("日期", 0) or 0)
        bd = dd.get("班段", {}) or {}
        recs = bd.get("入境11", []) or []
        if not isinstance(recs, list) or not recs:
            continue

        for r in recs:
            if not isinstance(r, dict):
                continue
            emp = str(r.get("原員工", r.get("人員", "")) or "").strip()
            if not emp:
                continue
            if _violates_in11_out_early(emp, d, people_dict):
                raise ValueError(f"[RULE] 入境11→隔天出境早班違規: {emp} on day {d}")


def _repair_in11_shortage(daily_list: list[dict], people_dict: dict, employee_cols) -> None:
    """Rescue module before scoring: fill 入境11 shortages via 入境10 swap."""
    base_cols = [c for c in employee_cols if c in people_dict]

    def _day_assigned_set(dd: dict) -> set[str]:
        assigned = set()
        for _, recs in (dd.get("班段", {}) or {}).items():
            if not isinstance(recs, list):
                continue
            for r in recs:
                if not isinstance(r, dict):
                    continue
                who = str(r.get("原員工", r.get("人員", "")) or "").strip()
                if who:
                    assigned.add(who)
        return assigned

    def _is_off(emp: str, d: int) -> bool:
        off_days = people_dict.get(emp, {}).get("休假", []) or []
        return d in off_days

    def _next_day_outbound_info(emp: str, d: int) -> tuple[bool, bool]:
        nd = _next_day_dict(d)
        if not nd or not isinstance(nd, dict):
            return (False, False)
        bd = nd.get("班段", {}) or {}
        has_outbound = any(sh in bd for sh in ("出境5", "出境6", "出境7", "出境8"))
        if not has_outbound:
            return (False, False)
        off_next = _is_off(emp, d + 1)
        pulled_next = False
        for sh in ("出境5", "出境6", "出境7", "出境8"):
            recs = bd.get(sh, []) or []
            if not isinstance(recs, list):
                continue
            for r in recs:
                if not isinstance(r, dict):
                    continue
                who = str(r.get("原員工", r.get("人員", "")) or "").strip()
                if who == emp and bool(r.get("拉班", False)):
                    pulled_next = True
                    break
            if pulled_next:
                break
        return (off_next, pulled_next)

    def _is_pull_for_shift(emp: str, dd: dict, shift_name: str) -> bool:
        need_team = _needed_team_for_shift(dd, shift_name)
        emp_team = str(people_dict.get(emp, {}).get("分組", "") or "")
        if need_team not in ("A", "B"):
            return False
        return emp_team != need_team

    for dd in (daily_list[:DAYS_LIMIT] or []):
        if not dd.get("班段"):
            continue
        if "入境11" not in dd.get("班段", {}):
            continue
        d = int(dd.get("日期", 0) or 0)
        need11 = int(shift_demands.get("入境11", 0) or 0)
        recs11 = dd.get("班段", {}).get("入境11", [])
        recs10 = dd.get("班段", {}).get("入境10", [])
        if not isinstance(recs11, list) or not isinstance(recs10, list):
            continue

        while len(recs11) < need11:
            assigned_today = _day_assigned_set(dd)
            A_candidates = [e for e in base_cols if (e not in assigned_today) and (not _is_off(e, d))]
            if not A_candidates:
                break

            A_chosen = None
            B_chosen = None
            B_rec = None

            for A in A_candidates:
                grpA = str(people_dict.get(A, {}).get("分組", "") or "")
                pool_same = [
                    r for r in recs10
                    if isinstance(r, dict)
                    and str(r.get("原員工", r.get("人員", "")) or "").strip() in people_dict
                    and str(r.get("cover", "") or "").strip() not in ("分隊長", "代理分隊長")
                    and str(people_dict.get(str(r.get("原員工", r.get("人員", "")) or "").strip(), {}).get("分組", "") or "") == grpA
                ]
                pool_other = [
                    r for r in recs10
                    if isinstance(r, dict)
                    and str(r.get("原員工", r.get("人員", "")) or "").strip() in people_dict
                    and str(r.get("cover", "") or "").strip() not in ("分隊長", "代理分隊長")
                    and str(people_dict.get(str(r.get("原員工", r.get("人員", "")) or "").strip(), {}).get("分組", "") or "") != grpA
                ]
                pool = pool_same if pool_same else pool_other
                if not pool:
                    continue

                # Priority: next-day outbound off, then next-day outbound pulled
                def _pri_key(rec: dict):
                    emp = str(rec.get("原員工", rec.get("人員", "")) or "").strip()
                    off_next, pulled_next = _next_day_outbound_info(emp, d)
                    if off_next:
                        return 0
                    if pulled_next:
                        return 1
                    return 2

                pool_sorted = sorted(pool, key=_pri_key)
                B_rec = pool_sorted[0]
                B_chosen = str(B_rec.get("原員工", B_rec.get("人員", "")) or "").strip()
                if B_chosen:
                    A_chosen = A
                    break

            if not A_chosen or not B_chosen or B_rec is None:
                break

            # remove B from 入境10
            try:
                recs10.remove(B_rec)
            except Exception:
                pass

            # adjust B old counts
            if bool(B_rec.get("拉班", False)):
                try:
                    people_dict[B_chosen]["拉班次數"] = max(0, int(people_dict[B_chosen].get("拉班次數", 0) or 0) - 1)
                except Exception:
                    pass
            old_cover = str(B_rec.get("cover", "") or "").strip()
            if old_cover not in ("", "填補", "分隊長", "代理分隊長"):
                fairness = people_dict.get(B_chosen, {}).get("公平性分數", {})
                if isinstance(fairness, dict):
                    try:
                        fairness[old_cover] = max(0, int(fairness.get(old_cover, 0) or 0) - 1)
                    except Exception:
                        pass
                    people_dict[B_chosen]["公平性分數"] = fairness
            try:
                counts = people_dict[B_chosen].get("班段次數", {})
                if isinstance(counts, dict):
                    counts["入境10"] = max(0, int(counts.get("入境10", 0) or 0) - 1)
                    people_dict[B_chosen]["班段次數"] = counts
            except Exception:
                pass

            # add A to 入境10
            A_pull = _is_pull_for_shift(A_chosen, dd, "入境10")
            recs10.append({
                "原員工": A_chosen,
                "人員": A_chosen,
                "代班人": "",
                "cover": "填補",
                "拉班": A_pull,
            })
            try:
                counts = people_dict[A_chosen].get("班段次數", {})
                if not isinstance(counts, dict):
                    counts = {}
                counts["入境10"] = int(counts.get("入境10", 0) or 0) + 1
                people_dict[A_chosen]["班段次數"] = counts
            except Exception:
                pass
            if A_pull:
                try:
                    people_dict[A_chosen]["拉班次數"] = int(people_dict[A_chosen].get("拉班次數", 0) or 0) + 1
                except Exception:
                    people_dict[A_chosen]["拉班次數"] = 1

            # add B to 入境11
            B_pull = _is_pull_for_shift(B_chosen, dd, "入境11")
            recs11.append({
                "原員工": B_chosen,
                "人員": B_chosen,
                "代班人": "",
                "cover": "填補",
                "拉班": B_pull,
            })
            try:
                counts = people_dict[B_chosen].get("班段次數", {})
                if not isinstance(counts, dict):
                    counts = {}
                counts["入境11"] = int(counts.get("入境11", 0) or 0) + 1
                people_dict[B_chosen]["班段次數"] = counts
            except Exception:
                pass
            if B_pull:
                try:
                    people_dict[B_chosen]["拉班次數"] = int(people_dict[B_chosen].get("拉班次數", 0) or 0) + 1
                except Exception:
                    people_dict[B_chosen]["拉班次數"] = 1


def _has_assignment_on_day(dd: dict, emp: str) -> bool:
    """True if emp is assigned to any shift on that day (by 原員工)."""
    bd = dd.get("班段", {}) or {}
    for _, recs in bd.items():
        if not isinstance(recs, list):
            continue
        for r in recs:
            if not isinstance(r, dict):
                continue
            who = str(r.get("原員工", r.get("人員", "")) or "").strip()
            if who == emp:
                return True
    return False


def _violates_no_empty_on_workday(
    daily_list: list[dict],
    employee_cols,
    people_dict: dict,
) -> bool:
    """If day is not 輪休, every base employee must be assigned or be off."""
    base_cols = [c for c in employee_cols if c in people_dict]
    for dd in (daily_list[:DAYS_LIMIT] or []):
        if not isinstance(dd, dict):
            continue
        # Skip 輪休 days (no shifts)
        if not dd.get("班段"):
            continue
        d = int(dd.get("日期", 0) or 0)
        for emp in base_cols:
            info = people_dict.get(emp, {}) if isinstance(people_dict, dict) else {}
            off_days = info.get("休假", []) or []
            if d in off_days:
                continue
            if _has_assignment_on_day(dd, emp):
                continue
            return True
    return False
            
def get_required_skills_for_shift(shift_name: str, skill_df: pd.DataFrame) -> list[str]:
    """\
    給定班段名稱（例如 '入境10' / '出境5' ），回傳該班段需要的職能列表。

    來源：skill 工作表的 A~G 欄（也就是 skill_df.columns[:7]）。
    - 需求欄：欄名若與 shift_name 相同，該欄的 1 表示「需要」
    - 職能名稱欄：優先找名為 '職位'/'職能'/'技能'，若找不到就用前 7 欄中最後一欄（通常是 G 欄）

    Return a list of required skills for the given shift.
    """

    # 只看 A~G（前 7 欄）
    front_cols = list(skill_df.columns[: min(7, len(skill_df.columns))])
    if not front_cols:
        return []

    # 1) 找需求欄（欄名 = shift_name）
    shift_name_norm = str(shift_name).strip()
    req_col = None
    for c in front_cols:
        if str(c).strip() == shift_name_norm:
            req_col = c
            break

    if req_col is None:
        # 找不到就回空（代表 skill 表裡沒有這個班段）
        return []

    # 2) 找職能名稱欄（你說通常是 G 欄，但做個保險）
    preferred_name_cols = ["職位", "職能", "技能", "能力"]
    name_col = None
    for c in front_cols:
        if str(c).strip() in preferred_name_cols:
            name_col = c
            break
    if name_col is None:
        name_col = front_cols[-1]  # fallback：當作 G 欄

    # 3) 掃每列：req_col == 1 => 加入 skill name
    required: list[str] = []

    for _, row in skill_df.iterrows():
        skill_name = str(row.get(name_col, "")).strip()
        if skill_name == "" or skill_name.lower() == "nan":
            continue

        v = row.get(req_col, None)
        if pd.isna(v):
            continue

        is_one = False
        try:
            is_one = int(v) == 1
        except Exception:
            is_one = str(v).strip() == "1"

        if is_one:
            required.append(skill_name)

    return required


# === 新增：assign_employees_to_shift helper ===
def assign_employees_to_shift(
    day_dict: dict,
    shift_name: str,
    required_skills: list[str],
    cand: list[str],
    Needed_Team: str,
    Demanded_Human_Resources: int,
    people_dict: dict,
) -> list[str]:
    """\
    指派某一天(day_dict)的某個班段(shift_name)人力。

    覆蓋式職能：required_skills 需被『整體覆蓋』，每個技能至少要有一位被指派的人具備。

    代班規則：
    - 若某員工在當天有「代班日期」：其職能視為抹掉（不能覆蓋 required_skills），但仍可用分組參與篩選。
    - 若該員工被選中排入班段，班段中記錄的是「代班人員」名字（抓不到就退回本人）。

    拉班（兩輪）：
    - 第 1 輪：只用 Needed_Team 排。
    - 第 2 輪：若第 1 輪排不滿需求，才用另一組（A<->B）拉班補人。
      * 拉班輪選人以「拉班次數」最低者優先；若本次有 target_skill，再用該技能公平性分數當次要排序；同分隨機。

    拉班次數更新（只在第 2 輪）：
    - 被拉班的人（chosen，從另一組被拉來上班）拉班次數 +1

    班段紀錄格式：
      {"人員": <name>, "cover": <skill-or-填補>, "拉班": <bool>}

    Return: 更新後的 cand（會把已選到的人從 cand 移除）。
    """

    # 保底：班段 key 必須存在
    if "班段" not in day_dict or shift_name not in day_dict.get("班段", {}):
        return cand

    # 需求人數
    try:
        need_n = int(Demanded_Human_Resources)
    except Exception:
        need_n = 0
    if need_n <= 0:
        return cand

    d = int(day_dict.get("日期", 0) or 0)

    # required_skills 正規化
    req_skills = [str(s).strip() for s in (required_skills or []) if str(s).strip()]

    # 只保留 cand 裡真的存在於 people_dict 的人
    cand_clean = [c for c in cand if c in people_dict]

    # ---- 代班輔助：判斷某員工當天是否有代班、以及代班人員名字 ----
    def _get_sub_helper(emp: str):
        info = people_dict.get(emp, {})

        # 優先用新結構：代班={日期: 代班人}
        sub_map = info.get("代班", {})
        if isinstance(sub_map, dict) and d in sub_map:
            return str(sub_map.get(d, "")).strip()

        # 相容舊結構（若還有殘留）
        sub_days = info.get("代班日期", []) or []
        sub_helpers = info.get("代班人員", []) or []

        # 轉成 int 清單
        sub_days_int = []
        for x in sub_days:
            try:
                sub_days_int.append(int(x))
            except Exception:
                try:
                    sub_days_int.append(int(str(x).strip()))
                except Exception:
                    pass

        if d not in sub_days_int:
            return None

        # 優先同 index
        try:
            idx = sub_days_int.index(d)
            if idx < len(sub_helpers):
                name = str(sub_helpers[idx]).strip()
                return name if name != "" else ""
        except Exception:
            pass

        # fallback：抓第一個非空代班人名
        for h in sub_helpers:
            hs = str(h).strip()
            if hs != "":
                return hs

        return ""  # 有代班但抓不到名字

    def _skills_for_today(emp: str) -> dict:
        # 若當天有代班 -> 職能抹掉
        if _get_sub_helper(emp) is not None:
            return {}
        return people_dict.get(emp, {}).get("職能", {}) or {}

    def _other_team(team: str) -> str:
        if team == "A":
            return "B"
        if team == "B":
            return "A"
        return ""

    def _pull_count(emp: str) -> int:
        try:
            return int(people_dict.get(emp, {}).get("拉班次數", 0) or 0)
        except Exception:
            return 0

    def _fairness(emp: str, sk: str) -> int:
        fairness = people_dict.get(emp, {}).get("公平性分數", {})
        if not isinstance(fairness, dict):
            return 0
        try:
            return int(fairness.get(sk, 0) or 0)
        except Exception:
            return 0

    def _shift_count(emp: str, sh: str) -> int:
        counts = people_dict.get(emp, {}).get("班段次數", {})
        if not isinstance(counts, dict):
            return 0
        try:
            return int(counts.get(sh, 0) or 0)
        except Exception:
            return 0

    def _pull_cap_threshold(pool: list[str]) -> int | None:
        """Soft cap to keep pull counts close (within +1 of current min if possible)."""
        if not pool:
            return None
        try:
            return min(_pull_count(e) for e in pool) + 1
        except Exception:
            return None

    def _had_inbound_cover_prev_day(emp: str) -> bool:
        """Soft preference: if emp did 入境 with cover=補入/值日 yesterday, prefer 出境8 today."""
        prev = _prev_day_dict(d)
        if not prev or not isinstance(prev, dict):
            return False
        bd = prev.get("班段", {}) or {}
        for sh in ("入境10", "入境11"):
            recs = bd.get(sh, []) or []
            if not isinstance(recs, list):
                continue
            for r in recs:
                if not isinstance(r, dict):
                    continue
                who = str(r.get("原員工", r.get("人員", "")) or "").strip()
                if who != emp:
                    continue
                cv = str(r.get("cover", "") or "").strip()
                if cv in ("補入", "值日"):
                    return True
        return False

    # -------------------------
    # Rule / Scorer pipeline (order = priority)
    # -------------------------
    def _rule_in11_no_out_early(eligible: list[str], ctx: dict) -> list[str]:
        if ctx.get("shift_name") != "入境11":
            return eligible
        return [e for e in eligible if not _violates_in11_out_early(e, d, people_dict)]

    def _sc_pull_over_cap(emp: str, ctx: dict) -> int:
        cap = ctx.get("cap")
        if cap is None:
            return 0
        return 1 if _pull_count(emp) > cap else 0

    def _sc_pull_count(emp: str, ctx: dict) -> int:
        return _pull_count(emp)

    def _sc_fairness(emp: str, ctx: dict) -> int:
        sk = ctx.get("target_skill", "")
        return _fairness(emp, sk) if sk else 0

    def _sc_inbound_prev_day_pref(emp: str, ctx: dict) -> int:
        if ctx.get("shift_name") != "出境8":
            return 0
        return 0 if _had_inbound_cover_prev_day(emp) else 1

    def _sc_shift_count(emp: str, ctx: dict) -> int:
        sh = ctx.get("shift_name", "")
        return _shift_count(emp, sh)

    def _apply_rules(eligible: list[str], rules: list, ctx: dict) -> list[str]:
        pool = eligible
        for rule in rules:
            pool = rule(pool, ctx)
            if not pool:
                break
        return pool

    def _score_tuple(emp: str, scorers: list, ctx: dict) -> tuple:
        return tuple(s(emp, ctx) for s in scorers)

    RULES_BASE = [_rule_in11_no_out_early]
    RULES_BY_TEAM = {
        "A": RULES_BASE,
        "B": RULES_BASE,
    }
    # Team-specific scorer priority (order = priority)
    TEAM_SCORERS_PULL = {
        "A": [_sc_pull_over_cap, _sc_pull_count, _sc_shift_count, _sc_inbound_prev_day_pref],
        "B": [_sc_pull_over_cap, _sc_pull_count, _sc_shift_count, _sc_inbound_prev_day_pref],
        "C": [_sc_pull_over_cap, _sc_pull_count, _sc_shift_count, _sc_inbound_prev_day_pref],
    }
    TEAM_SCORERS_NONPULL = {
        # A/B: fairness before shift count
        "A": [_sc_fairness, _sc_shift_count, _sc_inbound_prev_day_pref],
        "B": [_sc_fairness, _sc_shift_count, _sc_inbound_prev_day_pref],
        # C: shift count before fairness
        "C": [_sc_shift_count, _sc_fairness, _sc_inbound_prev_day_pref],
    }
    TEAM_PIPELINE = {
        "A": ["prepare", "rules", "scorers"],
        "B": ["prepare", "rules", "scorers"],
    }

    remaining_skills = req_skills.copy()

    # === helpers for Step 2/3: cover rebalance & skill-targeted replacement ===
    def _get_shift_recs() -> list[dict]:
        recs = day_dict.get("班段", {}).get(shift_name, [])
        return recs if isinstance(recs, list) else []

    def _emp_from_rec(rec: dict) -> str:
        return str(rec.get("原員工", rec.get("人員", "")) or "").strip()

    def _skill_covered_counts(recs: list[dict]) -> dict[str, int]:
        counts = {sk: 0 for sk in req_skills}
        for r in recs:
            emp = _emp_from_rec(r)
            if not emp:
                continue
            skills = _skills_for_today(emp)
            for sk in req_skills:
                if sk in skills:
                    counts[sk] += 1
        return counts

    def _rebalance_covers_within_shift() -> None:
        """Step 2: 在已選的人裡分配 cover 以覆蓋 required_skills。

        原則：
        - 先把所有非分隊長/代理分隊長的 cover 重置為『填補』（避免被 Step1 的 target_skill 汙染）。
        - 針對每個 required skill，挑一位具備該技能的人當 cover。
        - 盡量做到「一人一技能」；若人數不足才允許同一人被重複使用。
        - 被分配到 cover 的技能，公平性分數 +1。
        """
        recs = _get_shift_recs()
        if not recs or not req_skills:
            return

        # 1) reset covers (keep leader markers intact if any)
        for r in recs:
            cv = str(r.get("cover", "") or "").strip()
            if cv in ("分隊長", "代理分隊長"):
                continue
            r["cover"] = "填補"

        # 2) assign skills
        used_emp: set[str] = set()
        # 優先處理「越稀有」的技能
        counts = _skill_covered_counts(recs)
        skills_order = sorted(req_skills, key=lambda s: counts.get(s, 0))

        for sk in skills_order:
            # 找候選：先找未使用過的人（避免同人全包）
            candidates = []
            fallback = []
            for r in recs:
                emp = _emp_from_rec(r)
                if not emp:
                    continue
                if sk not in _skills_for_today(emp):
                    continue
                if emp not in used_emp:
                    candidates.append(emp)
                fallback.append(emp)

            pick_pool = candidates if candidates else fallback
            if not pick_pool:
                continue  # 這技能目前班段內沒人具備

            # 以技能公平性分數最低者優先（同分隨機）
            scores = {e: _fairness(e, sk) for e in pick_pool}
            m = min(scores.values())
            best = [e for e, sc in scores.items() if sc == m]
            chosen_emp = random.choice(best)

            # 寫回 rec
            for r in recs:
                if _emp_from_rec(r) == chosen_emp:
                    r["cover"] = sk
                    break

            # 更新公平性分數
            fairness = people_dict[chosen_emp].get("公平性分數", {})
            if not isinstance(fairness, dict):
                fairness = {}
            fairness[sk] = int(fairness.get(sk, 0) or 0) + 1
            people_dict[chosen_emp]["公平性分數"] = fairness

            used_emp.add(chosen_emp)

    def _find_replaceable_rec( ) -> dict | None:
        """找一個可以被替換掉的『填補』人員。

        原則：
        - 優先找 cover=填補 且其移除不會讓已覆蓋技能破洞的人。
        - 先嘗試找完全不具備任何 required skills 的人。
        """
        recs = _get_shift_recs()
        if not recs:
            return None

        # 先算每個技能目前被幾人覆蓋
        counts = _skill_covered_counts(recs)

        def _is_safe_remove(emp: str) -> bool:
            skills = _skills_for_today(emp)
            for sk in req_skills:
                if sk in skills and counts.get(sk, 0) <= 1:
                    return False
            return True

        fillers = []
        pure_fillers = []
        for r in recs:
            cv = str(r.get("cover", "") or "").strip()
            if cv != "填補":
                continue
            emp = _emp_from_rec(r)
            if not emp:
                continue
            if not _is_safe_remove(emp):
                continue
            fillers.append(r)
            # 完全沒有任何 required skills
            if all(sk not in _skills_for_today(emp) for sk in req_skills):
                pure_fillers.append(r)

        pool = pure_fillers if pure_fillers else fillers
        if not pool:
            return None

        # 可被替換的人中，優先選拉班次數低者被換掉（保護拉班次數高的人留在班段，避免反覆被丟回 cand 再被拉）
        scores = {id(r): _pull_count(_emp_from_rec(r)) for r in pool}
        mn = min(scores.values())
        best = [r for r in pool if scores.get(id(r), 0) == mn]
        return random.choice(best)

    def _swap_in_skill_person(missing_sk: str) -> bool:
        """Step 3: 若仍缺技能，從 cand 裡找具備 missing_sk 的人替換進來。

        - 先找 Needed_Team 的人（不算拉班）
        - 找不到才找另一組（算拉班，且拉班次數 +1）
        - 進來的人 cover 先設 missing_sk（之後 Step2 也會再整理一次）
        """
        rec_to_replace = _find_replaceable_rec()
        if rec_to_replace is None:
            return False

        # 候選池（仍在 cand_clean 內）
        same_team = []
        other_team = []
        for e in list(cand_clean):
            if missing_sk not in _skills_for_today(e):
                continue
            if shift_name == "入境11" and _violates_in11_out_early(e, d, people_dict):
                continue
            grp = people_dict.get(e, {}).get("分組", "")
            if grp == Needed_Team:
                same_team.append(e)
            elif grp == _other_team(Needed_Team):
                other_team.append(e)

        pick_pool = same_team if same_team else other_team
        if not pick_pool:
            return False

        # 選人：使用標準化 scoring pipeline（順序可調）
        if same_team:
            chosen_emp = _run_pipeline(Needed_Team, pick_pool, missing_sk, False)
            is_pull = False
        else:
            other_team = _other_team(Needed_Team)
            chosen_emp = _run_pipeline(other_team, pick_pool, missing_sk, True)
            is_pull = True
        if chosen_emp is None:
            return False

        # 把被替換的人退回 cand（讓他之後其他班段仍可用）
        replaced_emp = _emp_from_rec(rec_to_replace)
        if replaced_emp:
            if replaced_emp not in cand_clean and replaced_emp in people_dict:
                cand_clean.append(replaced_emp)
            if replaced_emp not in cand and replaced_emp in people_dict:
                cand.append(replaced_emp)

        # 從 shift recs 移除舊 rec
        recs = _get_shift_recs()
        try:
            recs.remove(rec_to_replace)
        except Exception:
            pass

        # 代班顯示
        sub_name = ""
        helper_name = _get_sub_helper(chosen_emp)
        if helper_name is not None:
            sub_name = str(helper_name).strip()

        # 新 rec
        recs.append({
            "原員工": chosen_emp,
            "人員": chosen_emp,
            "代班人": sub_name,
            "cover": missing_sk,
            "拉班": is_pull,
        })

        # 班段次數更新
        try:
            counts = people_dict[chosen_emp].get("班段次數", {})
            if not isinstance(counts, dict):
                counts = {}
            counts[shift_name] = int(counts.get(shift_name, 0) or 0) + 1
            people_dict[chosen_emp]["班段次數"] = counts
        except Exception:
            pass

        # 若為跨組拉班，拉班次數一定要 +1
        if is_pull:
            try:
                people_dict[chosen_emp]["拉班次數"] = int(people_dict[chosen_emp].get("拉班次數", 0) or 0) + 1
            except Exception:
                people_dict[chosen_emp]["拉班次數"] = 1

        # 從 cand 移除 chosen
        if chosen_emp in cand_clean:
            cand_clean.remove(chosen_emp)
        if chosen_emp in cand:
            cand.remove(chosen_emp)

        return True

    def _has_skill(emp: str, sk: str) -> bool:
        return sk in _skills_for_today(emp)

    def _choose_by_scoring(eligible: list[str], is_pull_round: bool, target_skill: str, team_to_use: str) -> str | None:
        if not eligible:
            return None
        ctx = {
            "shift_name": shift_name,
            "target_skill": target_skill,
            "is_pull_round": is_pull_round,
            "cap": _pull_cap_threshold(eligible) if is_pull_round else None,
            "team_to_use": team_to_use,
        }
        if is_pull_round:
            scorers = TEAM_SCORERS_PULL.get(team_to_use, TEAM_SCORERS_PULL["A"])
        else:
            if TEAM_SCORERS_NONPULL_OVERRIDE:
                order = TEAM_SCORERS_NONPULL_OVERRIDE.get(team_to_use, TEAM_SCORERS_NONPULL_OVERRIDE.get("A", []))
                order = order or ["fairness", "shift_count"]
                mapping = {
                    "fairness": _sc_fairness,
                    "shift_count": _sc_shift_count,
                }
                scorers = [mapping.get(k, _sc_fairness) for k in order] + [_sc_inbound_prev_day_pref]
            else:
                scorers = TEAM_SCORERS_NONPULL.get(team_to_use, TEAM_SCORERS_NONPULL["A"])
        scores = {emp: _score_tuple(emp, scorers, ctx) for emp in eligible}
        min_score = min(scores.values())
        best = [emp for emp, sc in scores.items() if sc == min_score]
        return random.choice(best)

    def _mod_prepare(ctx: dict) -> list[str]:
        team_ok = ctx.get("team_ok", []) or []
        target_skill = ctx.get("target_skill", "")
        is_pull_round = bool(ctx.get("is_pull_round", False))
        if target_skill:
            if is_pull_round:
                return team_ok[:]
            return [emp for emp in team_ok if _has_skill(emp, target_skill)]
        return team_ok[:]

    def _mod_rules(ctx: dict, eligible: list[str]) -> list[str]:
        rules = RULES_BY_TEAM.get(ctx.get("team_to_use", ""), RULES_BASE)
        return _apply_rules(eligible, rules, {"shift_name": shift_name})

    def _mod_scorers(ctx: dict, eligible: list[str]) -> str | None:
        return _choose_by_scoring(
            eligible,
            bool(ctx.get("is_pull_round", False)),
            ctx.get("target_skill", ""),
            ctx.get("team_to_use", ""),
        )

    PIPELINE_MODULES = {
        "prepare": _mod_prepare,
        "rules": _mod_rules,
        "scorers": _mod_scorers,
    }

    def _run_pipeline(team_to_use: str, team_ok: list[str], target_skill: str, is_pull_round: bool) -> str | None:
        if TEAM_PIPELINE_OVERRIDE:
            pipeline = TEAM_PIPELINE_OVERRIDE.get(team_to_use, TEAM_PIPELINE_OVERRIDE.get("A", ["prepare", "rules", "scorers"]))
        else:
            pipeline = TEAM_PIPELINE.get(team_to_use, ["prepare", "rules", "scorers"])
        ctx = {
            "team_to_use": team_to_use,
            "team_ok": team_ok,
            "target_skill": target_skill,
            "is_pull_round": is_pull_round,
        }
        eligible: list[str] = team_ok[:]
        chosen: str | None = None
        for name in pipeline:
            mod = PIPELINE_MODULES.get(name)
            if not mod:
                continue
            if name == "prepare":
                eligible = mod(ctx)
            elif name == "rules":
                eligible = mod(ctx, eligible)
            elif name == "scorers":
                chosen = mod(ctx, eligible)
            if name != "scorers" and not eligible:
                return None
        return chosen


    # 單輪執行（team_to_use / 是否拉班輪）
    def _run_round(team_to_use: str, slots: int, is_pull_round: bool) -> int:
        nonlocal cand_clean, remaining_skills

        added = 0
        for _ in range(slots):
            team_ok = [emp for emp in cand_clean if people_dict.get(emp, {}).get("分組", "") == team_to_use]
            if not team_ok:
                break

            # 決定本次目標技能（每次最多覆蓋一個）
            target_skill = ""
            if remaining_skills:
                counts = {sk: sum(1 for emp in team_ok if _has_skill(emp, sk)) for sk in remaining_skills}
                viable = [sk for sk in remaining_skills if counts.get(sk, 0) > 0]
                if viable:
                    target_skill = min(viable, key=lambda s: counts.get(s, 0))
            chosen = _run_pipeline(team_to_use, team_ok, target_skill, is_pull_round)
            if chosen is None:
                break

            # 代班顯示規則：
            # - 班段紀錄的「原員工」永遠是 chosen（請代班者 / 被排入該班段者）
            # - 若當天有代班，額外記錄「代班人」= 代班姓名（可能不在 people_dict）
            sub_name = ""
            helper_name = _get_sub_helper(chosen)
            if helper_name is not None:
                sub_name = str(helper_name).strip()  # helper_name 可能是 "" 代表填不出名字

            # 只有真的具備 target_skill 才算覆蓋，否則就是填補
            cover_role = target_skill if (target_skill and _has_skill(chosen, target_skill)) else "填補"
            day_dict["班段"][shift_name].append({
                "原員工": chosen,
                "人員": chosen,
                "代班人": sub_name,
                "cover": cover_role,
                "拉班": is_pull_round,
            })

            # 班段次數更新
            try:
                counts = people_dict[chosen].get("班段次數", {})
                if not isinstance(counts, dict):
                    counts = {}
                counts[shift_name] = int(counts.get(shift_name, 0) or 0) + 1
                people_dict[chosen]["班段次數"] = counts
            except Exception:
                pass

            # 公平性分數：只對本次真正覆蓋的技能 +1
            if cover_role != "填補":
                fairness = people_dict[chosen].get("公平性分數", {})
                if not isinstance(fairness, dict):
                    fairness = {}
                fairness[cover_role] = int(fairness.get(cover_role, 0) or 0) + 1
                people_dict[chosen]["公平性分數"] = fairness

                if cover_role in remaining_skills:
                    remaining_skills.remove(cover_role)

            # 拉班次數更新（只在拉班輪）
            if is_pull_round:
                # 被拉班的人一定要補償：拉班次數 +1
                try:
                    people_dict[chosen]["拉班次數"] = int(people_dict[chosen].get("拉班次數", 0) or 0) + 1
                except Exception:
                    people_dict[chosen]["拉班次數"] = 1

            # 從候選移除（移除 chosen 本人）
            if chosen in cand_clean:
                cand_clean.remove(chosen)
            if chosen in cand:
                cand.remove(chosen)

            added += 1

        return added

    def _run_round_any(slots: int) -> int:
        """Rescue round: pull from any remaining candidates to fill empty slots."""
        nonlocal cand_clean, remaining_skills
        added = 0
        for _ in range(slots):
            team_ok = cand_clean[:]
            if not team_ok:
                break
            # no skill targeting in rescue round
            chosen = _run_pipeline("A", team_ok, "", True)
            if chosen is None:
                break
            sub_name = ""
            helper_name = _get_sub_helper(chosen)
            if helper_name is not None:
                sub_name = str(helper_name).strip()

            day_dict["班段"][shift_name].append({
                "原員工": chosen,
                "人員": chosen,
                "代班人": sub_name,
                "cover": "填補",
                "拉班": True,
            })

            # 班段次數更新
            try:
                counts = people_dict[chosen].get("班段次數", {})
                if not isinstance(counts, dict):
                    counts = {}
                counts[shift_name] = int(counts.get(shift_name, 0) or 0) + 1
                people_dict[chosen]["班段次數"] = counts
            except Exception:
                pass

            # 拉班次數更新
            try:
                people_dict[chosen]["拉班次數"] = int(people_dict[chosen].get("拉班次數", 0) or 0) + 1
            except Exception:
                people_dict[chosen]["拉班次數"] = 1

            if chosen in cand_clean:
                cand_clean.remove(chosen)
            if chosen in cand:
                cand.remove(chosen)

            added += 1

        return added

    # ===== Round 1：正常排 =====
    # Soft preference: if 出境7+出境8 combined need exceeds same-team supply,
    # pull into 出境7 first (leave same-team for 出境8) when possible.
    round1_slots = need_n
    if shift_name == "出境7":
        same_team_cnt = sum(1 for e in cand_clean if people_dict.get(e, {}).get("分組", "") == Needed_Team)
        other_team = _other_team(Needed_Team)
        other_team_cnt = sum(1 for e in cand_clean if people_dict.get(e, {}).get("分組", "") == other_team) if other_team else 0
        demand8_total = int(shift_demands.get("出境8", 0) or 0)
        already8 = len(day_dict.get("班段", {}).get("出境8", []) or [])
        need8 = max(0, demand8_total - already8)
        cap_same_for_7 = max(0, same_team_cnt - need8)
        if cap_same_for_7 < need_n:
            if other_team_cnt >= (need_n - cap_same_for_7):
                round1_slots = cap_same_for_7

    added_1 = _run_round(Needed_Team, round1_slots, is_pull_round=False)

    # ===== Round 2：不夠才拉班 =====
    remaining_slots = need_n - added_1
    if remaining_slots > 0:
        other = _other_team(Needed_Team)
        if other:
            _run_round(other, remaining_slots, is_pull_round=True)

    # ===== Round 3：救援拉班（全組別，直到補滿或人用完）=====
    remaining_slots = need_n - len(day_dict["班段"][shift_name])
    if remaining_slots > 0 and RESCUE_FILL_ALL:
        _run_round_any(remaining_slots)

    # ===== Step 2：班段內重新分配 cover，盡量用現有人員完成技能覆蓋 =====
    _rebalance_covers_within_shift()

    # ===== Step 3：若仍缺技能，才做技能導向的「替換式拉班」 =====
    # 計算目前仍未被覆蓋的技能
    recs_now = _get_shift_recs()
    covered_now = {sk: False for sk in req_skills}
    for r in recs_now:
        emp = _emp_from_rec(r)
        if not emp:
            continue
        skills = _skills_for_today(emp)
        for sk in req_skills:
            if sk in skills:
                covered_now[sk] = True

    missing = [sk for sk, ok in covered_now.items() if not ok]

    # 針對缺少技能逐一嘗試替換補洞，成功後再做一次 cover 整理
    changed = False
    for sk in missing:
        if _swap_in_skill_person(sk):
            changed = True

    if changed:
        _rebalance_covers_within_shift()

    return cand


# === 新增：分隊長請假時，安排「代理分隊長」並從 cand 除名 ===
def reserve_acting_leader_if_needed(
    day_dict: dict,
    cand: list[str],
    people_dict: dict,
) -> list[str]:
    """\
    規則：若當天「分隊長」全部都在休假（或當天無班），
    則必須從 cand 中挑一位具備「代理分隊長」職能的人來代理。

    - 代理分隊長當天的上勤時間視為與分隊長相同（此版先用資料欄位標記，
      之後你在出飛/時間層再真正套用時段）。
    - 代理分隊長因為有職務在身，當天會從 cand 移除（不能再被排到其他位置）。

    回傳：更新後的 cand

    English:
    If no team leader (分隊長) is available today, reserve an acting leader (代理分隊長) from candidates and remove them from cand.
    """

    # 當天若沒有任何班段（例如輪休），就不用處理
    if not day_dict.get("班段"):
        return cand

    d = int(day_dict.get("日期", 0) or 0)

    # 1) 判斷當天是否「分隊長全請假」
    leader_available = False
    for emp, info in people_dict.items():
        # C 代表分隊長；另外也容許用「分隊長」職能標記
        skills = info.get("職能", {})
        is_leader = (info.get("分組", "") == "C") or ("分隊長" in skills)
        if not is_leader:
            continue
        off_days = info.get("休假", [])
        if d not in off_days:
            leader_available = True
            break

    if leader_available:
        return cand

    # 2) 從 cand 中找具備「代理分隊長」的人（只保留真的存在於 people_dict 的人）
    cand_clean = [c for c in cand if c in people_dict]
    acting_pool = [
        emp for emp in cand_clean
        if "代理分隊長" in people_dict.get(emp, {}).get("職能", {})
    ]

    if not acting_pool:
        # 依規則不應該發生：若分隊長全請假，必須有人可代理
        raise ValueError(
            f"[ERROR] {d}號：分隊長全請假，但 cand 中找不到具備『代理分隊長』職能的人可代理。"
        )

    # 3) 挑公平性分數最低者（同分隨機）
    def _acting_score(emp: str) -> int:
        fairness = people_dict.get(emp, {}).get("公平性分數", {})
        if not isinstance(fairness, dict):
            return 0
        try:
            return int(fairness.get("代理分隊長", 0) or 0)
        except Exception:
            return 0

    scores = {emp: _acting_score(emp) for emp in acting_pool}
    min_score = min(scores.values())
    best = [emp for emp, sc in scores.items() if sc == min_score]
    chosen = random.choice(best)

    # 4) 記錄到 day_dict（先用欄位標記，後面時間層再真正套用）
    day_dict.setdefault("特殊職務", {})
    day_dict["特殊職務"]["分隊長"] = chosen
    day_dict["特殊職務"]["分隊長類型"] = "代理"

    # 5) 更新公平性分數：代理分隊長 +1
    fairness = people_dict[chosen].get("公平性分數", {})
    if not isinstance(fairness, dict):
        fairness = {}
    fairness["代理分隊長"] = int(fairness.get("代理分隊長", 0) or 0) + 1
    people_dict[chosen]["公平性分數"] = fairness

    # 6) 從 cand 移除（不能再排其他職位）
    if chosen in cand:
        cand.remove(chosen)

    return cand


# === 新增：強制分隊長(C)進入指定班段（入境10 / 出境6），並從 cand 除名 ===
def reserve_leader_for_mandatory_shifts(
    day_dict: dict,
    cand: list[str],
    people_dict: dict,
) -> list[str]:
    """\
    規則：C 代表分隊長。
    - 入境日：分隊長一定在「入境10」
    - 出境日：分隊長一定在「出境6」
    - 若當天由代理分隊長代理，則代理者同樣適用此規則。

    行為：
    - 把分隊長/代理分隊長直接塞進對應班段 list
    - 從 cand 移除（當天不能再排其他職位）
    - 更新公平性分數（分隊長 或 代理分隊長）+1

    回傳：更新後的 cand。
    """

    if not day_dict.get("班段"):
        return cand

    d = int(day_dict.get("日期", 0) or 0)

    # 1) 先決定當天的 leader 是誰：
    #    - 若前面已經安排代理：用代理者
    #    - 否則：找分組為 C 且未休假的人（通常只有一位）
    leader_type = "正式"
    leader_emp = None

    if "特殊職務" in day_dict and isinstance(day_dict["特殊職務"], dict):
        if day_dict["特殊職務"].get("分隊長類型") == "代理" and day_dict["特殊職務"].get("分隊長"):
            leader_emp = day_dict["特殊職務"].get("分隊長")
            leader_type = "代理"

    if leader_emp is None:
        for emp, info in people_dict.items():
            if info.get("分組", "") != "C":
                continue
            off_days = info.get("休假", [])
            if d in off_days:
                continue
            leader_emp = emp
            break

    if not leader_emp or leader_emp not in people_dict:
        # 依規則不應該發生：必定有分隊長或代理分隊長可用
        raise ValueError(
            f"[ERROR] {d}號：找不到當日分隊長（C）或代理分隊長，無法強制指派入境10/出境6。"
        )

    # 2) 依 day_dict 的班段型態強制塞入
    if "入境10" in day_dict["班段"]:
        must_shift = "入境10"
    elif "出境6" in day_dict["班段"]:
        must_shift = "出境6"
    else:
        return cand

    if leader_emp not in day_dict["班段"][must_shift]:
        cover = "代理分隊長" if leader_type == "代理" else "分隊長"
        day_dict["班段"][must_shift].append({
            "原員工": leader_emp,
            "人員": leader_emp,
            "代班人": "",
            "cover": cover,
            "拉班": False,
        })

    # 3) 記錄到 day_dict（方便之後時間層套用）
    day_dict.setdefault("特殊職務", {})
    day_dict["特殊職務"]["當日分隊長"] = leader_emp
    day_dict["特殊職務"]["當日分隊長類型"] = leader_type

    # 4) 更新公平性分數
    fairness_key = "代理分隊長" if leader_type == "代理" else "分隊長"
    fairness = people_dict[leader_emp].get("公平性分數", {})
    if not isinstance(fairness, dict):
        fairness = {}
    fairness[fairness_key] = int(fairness.get(fairness_key, 0) or 0) + 1
    people_dict[leader_emp]["公平性分數"] = fairness

    # 4.1) 更新班段次數
    try:
        counts = people_dict[leader_emp].get("班段次數", {})
        if not isinstance(counts, dict):
            counts = {}
        counts[must_shift] = int(counts.get(must_shift, 0) or 0) + 1
        people_dict[leader_emp]["班段次數"] = counts
    except Exception:
        pass

    # 5) 從 cand 移除
    if leader_emp in cand:
        cand.remove(leader_emp)

    return cand

# === 參數設定：取得各時段需求人數 ===
# 你的「參數設定」工作表是「橫向」：
# - 第 1 列是欄位名稱（例如 入境10人數、出境8人數...）
# - 第 2 列是對應數值
# 所以直接用 variables.iloc[0]（第一筆資料列）去拿每個欄位即可。

# print("[DEBUG] variables columns:", list(variables.columns))
# print("[DEBUG] variables head(3):\n", variables.head(3))


def get_shift_demands(variables_df: pd.DataFrame) -> tuple[str, dict[str, int]]:
    """\
    回傳 (第一輪早班代號, 各班段需求人數 dict)

    Example return:
      ("A", {"入境10": 6, "入境11": 6, "出境5": 3, ...})
    """

    if variables_df is None or len(variables_df) == 0:
        return "", {}

    row = variables_df.iloc[0]

    # 第一輪早班？（A/B）
    first_team = str(row.get("第一輪早班？", "")).strip()

    # 欄位名稱 -> 班段名稱
    col_to_shift = {
        "入境10人數": "入境10",
        "入境11人數": "入境11",
        "出境5人數": "出境5",
        "出境6人數": "出境6",
        "出境7人數": "出境7",
        "出境8人數": "出境8",
    }

    demands: dict[str, int] = {}
    for col, shift in col_to_shift.items():
        v = row.get(col, 0)
        num = pd.to_numeric(v, errors="coerce")
        demands[shift] = int(num) if not pd.isna(num) else 0

    return first_team, demands




# print("[DEBUG] shift_demands=", shift_demands)
# 快速驗證：印前兩天的空殼（之後你要開始塞資料再把這段註解打開即可）
# print("[DEBUG] first 2 day skeletons:")
# === Level 1 排班順序調整 ===
# 原本想用 round 控制「第一輪排出境、第二輪排入境」，
# 但因為「入境晚班不能接出境早班（除非其中一天是請代班）」
# 所以改成全月（或整段範圍）依序：
# 1) 出境早班（出境5、出境6）
# 2) 入境早班（入境10）
# 3) 入境晚班（入境11）
# 4) 出境晚班（出境7、出境8）


# =========================
# Multi-run search (minimal intrusion):
# Repeat scheduling and keep the roster with the best (lowest) pull-count std.
# Stop when the current best hasn't improved for SEARCH_PATIENCE tries.
# =========================
SEARCH_BEST_ROSTER = True
SEARCH_MAX_TRIES = 5000   # hard cap to avoid infinite loops
SEARCH_MIN_TRIES = 100    # always try at least this many times before early-stop
SEARCH_PATIENCE = 10      # stop if no improvement within this many tries after best



def _pull_std_ab(people_dict: dict) -> float:
    """Standard deviation of pull counts across A/B employees (lower is better)."""
    pulls: list[int] = []
    for _, info in (people_dict or {}).items():
        if not isinstance(info, dict):
            continue
        grp = str(info.get("分組", "") or "").strip().upper()
        if grp not in ("A", "B"):
            continue
        try:
            pulls.append(int(info.get("拉班次數", 0) or 0))
        except Exception:
            pulls.append(0)

    if len(pulls) <= 1:
        return 0.0
    return float(statistics.pstdev(pulls))


def _fairness_sum_std_ab(people_dict: dict) -> float:
    """Standard deviation of total fairness scores across A/B employees (lower is better)."""
    vals: list[int] = []
    for _, info in (people_dict or {}).items():
        if not isinstance(info, dict):
            continue
        grp = str(info.get("分組", "") or "").strip().upper()
        if grp not in ("A", "B"):
            continue
        fairness = info.get("公平性分數", {})
        if not isinstance(fairness, dict):
            vals.append(0)
            continue
        total = 0
        for v in fairness.values():
            try:
                total += int(v or 0)
            except Exception:
                continue
        vals.append(total)

    if len(vals) <= 1:
        return 0.0
    return float(statistics.pstdev(vals))


def _shift_count_std_ab(people_dict: dict) -> float:
    """Sum of std devs across shifts for A/B employees (lower is better)."""
    total = 0.0
    for sh in SHIFT_ORDER:
        vals: list[int] = []
        for _, info in (people_dict or {}).items():
            if not isinstance(info, dict):
                continue
            grp = str(info.get("分組", "") or "").strip().upper()
            if grp not in ("A", "B"):
                continue
            counts = info.get("班段次數", {})
            if not isinstance(counts, dict):
                vals.append(0)
                continue
            try:
                vals.append(int(counts.get(sh, 0) or 0))
            except Exception:
                vals.append(0)
        if len(vals) <= 1:
            continue
        total += float(statistics.pstdev(vals))
    return total

# New helper: check all A/B 拉班次數 >= 1
def _all_pulls_nonzero_ab(people_dict: dict) -> bool:
    """Return True only if every A/B employee has pull-count >= 1."""
    for _, info in (people_dict or {}).items():
        if not isinstance(info, dict):
            continue
        grp = str(info.get("分組", "") or "").strip().upper()
        if grp not in ("A", "B"):
            continue
        try:
            pulls = int(info.get("拉班次數", 0) or 0)
        except Exception:
            pulls = 0
        if pulls <= 0:
            return False
    return True


def _needed_team_for_shift(dd: dict, shift_name: str) -> str:
    """依班段決定所需組別（含智能選隊，降低硬拉班）。

    先照你原本輪替直覺決定 preferred：
    - 出境5/6：用早班的另一組
    - 入境11：用早班的另一組
    - 其他：用早班

    """

    base = dd.get("早班", "")
    if base not in ("A", "B"):
        return base

    other = "B" if base == "A" else "A"

    # 1) your original rule -> preferred
    preferred = other if shift_name in ("出境5", "出境6", "入境11") else base

    return preferred


def _assign_one_shift(dd: dict, shift_name: str) -> None:
    """對單一日的單一班段進行指派，會更新 dd["_cand"] 與 dd["班段"][shift_name]。"""
    if shift_name not in dd.get("班段", {}):
        return

    Needed_Team = _needed_team_for_shift(dd, shift_name)
    required_skills = get_required_skills_for_shift(shift_name, skill)
    # Debug: team composition inside today's remaining candidate pool
    if DEBUG_SCHED:
        cand_now = dd.get("_cand", []) or []
        other_team = "B" if Needed_Team == "A" else ("A" if Needed_Team == "B" else "")
        team_cnt = sum(1 for e in cand_now if people_dict.get(e, {}).get("分組", "") == Needed_Team)
        other_cnt = sum(1 for e in cand_now if people_dict.get(e, {}).get("分組", "") == other_team) if other_team else 0
        _debug(
            f"[POOL ] day {dd['日期']} before {shift_name}: cand={len(cand_now)} "
            f"{Needed_Team}={team_cnt} {other_team}={other_cnt} base早班={dd.get('早班','')}"
        )
    # 需求人數扣掉已經預先塞入的人（例如分隊長強制進班段）
    demand_total = int(shift_demands.get(shift_name, 0) or 0)
    Demanded_Human_Resources = max(0, demand_total - len(dd["班段"][shift_name]))

    dd["_cand"] = assign_employees_to_shift(
        day_dict=dd,
        shift_name=shift_name,
        required_skills=required_skills,
        cand=(dd.get("_cand", []) or []),
        Needed_Team=Needed_Team,
        Demanded_Human_Resources=Demanded_Human_Resources,
        people_dict=people_dict,
    )

    # Debug: show cand shrink, assigned, and pulled info
    if DEBUG_SCHED:
        assigned = dd["班段"][shift_name]
        pulled = [r for r in assigned if isinstance(r, dict) and r.get("拉班")]
        _debug(
            f"[ASSIGN] {dd['日期']} {shift_name} team={Needed_Team} "
            f"need={Demanded_Human_Resources} req_skills={required_skills} "
            f"assigned={len(assigned)} pulled={len(pulled)} cand_left={len(dd.get('_cand', []) or [])}"
        )
        if pulled:
            # show a compact list of pulled employees (原員工)
            pulled_names = [str(r.get('原員工', '')) for r in pulled]
            _debug("        pulled: " + ", ".join(pulled_names))


def _schedule_once() -> None:
    """Run the whole scheduling once (mutates daily_list[:DAYS_LIMIT] and people_dict)."""

    global DAYS_LIMIT

    # 0) Reset mutable state when re-running in batches (VERY IMPORTANT)
    if RESET_BEFORE_SCHED:
        _debug("[RESET] Resetting schedule state (daily_list + people_dict)...")
        reset_schedule_state(daily_list[:DAYS_LIMIT], people_dict)

    # 1) Initialize per-day candidate lists (always rebuild after reset)
    for dd in daily_list[:DAYS_LIMIT]:
        # 只處理有班段的日子（輪休 dd["班段"] 會是 {}）
        if not dd.get("班段"):
            dd["_cand"] = []
            continue

        cand = get_candidates_for_day(dd, employee_cols, people_dict)

        # 若分隊長當天請假：先安排代理分隊長，並從 cand 除名
        cand = reserve_acting_leader_if_needed(dd, cand, people_dict)

        # C 代表分隊長：強制分隊長/代理分隊長進入指定班段（入境10 / 出境6），並從 cand 除名
        cand = reserve_leader_for_mandatory_shifts(dd, cand, people_dict)

        dd["_cand"] = cand

    # Optional: snapshot to verify reset worked
    debug_state_snapshot(daily_list[:DAYS_LIMIT], people_dict, days_n=3)

    # 2) 出境早班：先排出境5、出境6
    for dd in daily_list[:DAYS_LIMIT]:
        if not dd.get("班段"):
            continue
        _assign_one_shift(dd, "出境5")
        _assign_one_shift(dd, "出境6")

    # 3) 入境早班：排入境10
    for dd in daily_list[:DAYS_LIMIT]:
        if not dd.get("班段"):
            continue
        _assign_one_shift(dd, "入境10")

    # 4) 入境晚班：排入境11
    for dd in daily_list[:DAYS_LIMIT]:
        if not dd.get("班段"):
            continue
        _assign_one_shift(dd, "入境11")

    # 5) 出境晚班：最後排出境7、出境8
    for dd in daily_list[:DAYS_LIMIT]:
        if not dd.get("班段"):
            continue
        _assign_one_shift(dd, "出境7")
        _assign_one_shift(dd, "出境8")

    _validate_hard_rules_in11_out_early(daily_list[:DAYS_LIMIT], people_dict)
    
    # Final snapshot after scheduling
    debug_state_snapshot(daily_list[:DAYS_LIMIT], people_dict, days_n=5)



def _shift_cell_text(shift_name: str, rec: dict) -> str:
    """把班段 + cover + 拉班轉成班表格子內的短字串。"""
    # 取數字：入境10 -> 10, 出境6 -> 6
    m = re.search(r"(\d+)", str(shift_name))
    num = m.group(1) if m else str(shift_name)

    cover = str(rec.get("cover", "") or "").strip()
    pulled = bool(rec.get("拉班", False))

    # cover 簡寫（你可依實務擴充）
    cover_map = {
        "入境公": "公",
        "出境早公": "公",
        "出境晚公": "公",
        "值日": "值",
        "補入": "補",
        "分隊長": "",
        "代理分隊長": "代理",
    }

    suffix = ""
    if cover and cover != "填補":
        suffix = cover_map.get(cover, cover[:1])  # fallback：取第一個字

    pull_mark = ""

    # 若有代班人：直接顯示「代班人+班號」（符合你要的：例如 秋芳10）
    sub = str(rec.get("代班人", "") or "").strip()
    if sub != "":
        return f"{sub}{num}{pull_mark}"

    return f"{num}{suffix}{pull_mark}"


def build_roster_table(
    daily_list: list[dict],
    employee_cols,
    people_dict: dict,
    include_external: bool = True
) -> pd.DataFrame:
    """
    產生排班表 DataFrame：index=日期, columns=人員, value=班型
    include_external=True：若班段中出現 people_dict 以外的人名（外援/代班人），也會加成額外欄位
    """

    # 1) 日期排序
    days_sorted = [dd for dd in daily_list if isinstance(dd, dict) and "日期" in dd]
    days_sorted.sort(key=lambda x: int(x.get("日期", 0) or 0))
    date_index = [int(dd.get("日期", 0) or 0) for dd in days_sorted]

    # 2) 欄位：people_dict 員工（依 employee_cols 順序）
    base_cols = [c for c in employee_cols if c in people_dict]

    # 3) 額外欄位：外援/代班人（若你想把秋芳、品心、坤政這種也顯示）
    extra_cols: list[str] = []
    if include_external:
        seen = set(base_cols)
        for dd in days_sorted:
            for sh, recs in (dd.get("班段", {}) or {}).items():
                if not isinstance(recs, list):
                    continue
                for rec in recs:
                    if not isinstance(rec, dict):
                        continue
                    name = str(rec.get("人員", "") or "").strip()
                    if name and name not in seen:
                        extra_cols.append(name)
                        seen.add(name)

    cols = base_cols + extra_cols

    # 4) 建表（先塞空字串）
    df = pd.DataFrame("", index=date_index, columns=cols)
    df.index.name = "日期"

    # 5) 填值：同一天同一人若被排到兩個班段，用 '/' 串起來
    for dd in days_sorted:
        d = int(dd.get("日期", 0) or 0)
        for sh, recs in (dd.get("班段", {}) or {}).items():
            if not isinstance(recs, list):
                continue
            for rec in recs:
                if not isinstance(rec, dict):
                    continue
                # 以「原員工」為準：代班也填回請代班者的欄位
                name = str(rec.get("原員工", rec.get("人員", "")) or "").strip()
                if not name:
                    continue

                if name not in df.columns:
                    # include_external=False 時會走到這裡，直接略過外援
                    continue

                cell = _shift_cell_text(sh, rec)
                prev = df.at[d, name]
                df.at[d, name] = cell if prev == "" else f"{prev}/{cell}"

    # 6) 填入休假標記（僅限本部人員）
    for emp in base_cols:
        info = people_dict.get(emp, {})
        off_days = info.get("休假", []) or []
        for od in off_days:
            try:
                day = int(od)
            except Exception:
                continue
            if day in df.index and df.at[day, emp] == "":
                df.at[day, emp] = "休"

    return df



if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Airport roster generator (web-friendly)")
    parser.add_argument("input", help="Input Excel template path (.xlsx)")
    parser.add_argument("-o", "--output", default="", help="Output Excel path (.xlsx)")
    parser.add_argument("--days", type=int, default=28, help="How many days to schedule")
    parser.add_argument("--no-search", action="store_true", help="Disable multi-try search")
    parser.add_argument("--tries", type=int, default=5000, help="Max tries when searching")
    parser.add_argument("--patience", type=int, default=10, help="Stop after N non-improving tries")
    parser.add_argument("--allow-zero-pulls", action="store_true", help="Allow some A/B pull-count to be 0")
    parser.add_argument("--debug", action="store_true", help="Enable debug prints")
    parser.add_argument("--seed", type=int, default=None, help="Random seed")

    args = parser.parse_args()

    result = run_scheduler(
        input_excel_path=args.input,
        output_excel_path=(args.output if args.output else None),
        days_limit=args.days,
        search_best_roster=(not args.no_search),
        search_max_tries=args.tries,
        search_patience=args.patience,
        require_all_pulls_nonzero=(not args.allow_zero_pulls),
        debug=args.debug,
        random_seed=args.seed,
    )

    print(
        f"[OK] output={result['output_path']} tries={result['tries']} "
        f"best_score={result['best_std']:.4f} pull_std={result['best_pull_std']:.4f} "
        f"fair_std={result['best_fair_std']:.4f} used_search={result['used_search']}"
    )
