from __future__ import annotations

import shutil
import tempfile
from pathlib import Path
from functools import lru_cache
import importlib
import threading
import time
import uuid
import html
from openpyxl import load_workbook
import pandas as pd

from fastapi import FastAPI, UploadFile, File, HTTPException, Form, BackgroundTasks, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 讓 Python 找得到同層/上層的「機場排班程式.py」
import sys
BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(PROJECT_DIR))


_SCHEDULE_MODES: dict[str, dict] = {
    "monthly": {
        "label": "月班表",
        "desc": "一個月前排下個月班表，重視整月公平與穩定。",
        "run_kwargs": {
            "search_best_roster": True,
            "search_patience": 10,
            "require_all_pulls_nonzero": False,
            "rescue_fill": True,
        },
    },
    "departure": {
        "label": "出境勤務表",
        "desc": "前一天排隔天出境勤務表，重視即時可用與快速收斂。",
        "run_kwargs": {
            "search_best_roster": True,
            "search_patience": 5,
            "require_all_pulls_nonzero": False,
            "rescue_fill": True,
        },
    },
}


def _mode_key_or_default(mode: str | None) -> str:
    key = str(mode or "").strip().lower()
    return key if key in _SCHEDULE_MODES else "monthly"


@lru_cache(maxsize=1)
def _get_scheduler_funcs():
    import 機場排班程式 as monthly_mod  # noqa: E402
    departure_mod = None
    departure_api = "none"  # none | new | legacy
    departure_import_error: Exception | None = None
    try:
        importlib.invalidate_caches()
        departure_mod = importlib.import_module("departure_duty_scheduler")
        departure_api = "new"
    except Exception as e:
        departure_import_error = e
        try:
            departure_mod = importlib.import_module("departure_scheduler")
            departure_api = "legacy"
        except Exception:
            pass

    run_monthly = getattr(monthly_mod, "run_scheduler")
    validate_monthly = getattr(monthly_mod, "validate_input_excel")

    def validate_departure(input_excel_path: str) -> list[dict]:
        if departure_api == "none":
            reason = (
                "departure scheduler module not found on server. "
                "請確認已部署 `departure_duty_scheduler.py`（或舊版 `departure_scheduler.py`）。"
            )
            if departure_import_error is not None:
                reason = f"{reason} 原始錯誤: {departure_import_error}"
            return [
                {
                    "sheet": "Departure",
                    "columns": [],
                    "reason": reason,
                }
            ]
        try:
            assert departure_mod is not None
            if departure_api == "new":
                emp_df, dem_df = departure_mod.read_input(input_excel_path)
                departure_mod.validate_input(emp_df, dem_df)
            else:
                # Legacy departure wrapper (monthly-engine based)
                validate_fn = getattr(departure_mod, "validate_departure_input_excel", None)
                if callable(validate_fn):
                    return validate_fn(input_excel_path) or []
                raise RuntimeError("legacy departure module 缺少 validate_departure_input_excel")
            return []
        except Exception as e:
            return [{"sheet": "Departure", "columns": [], "reason": str(e)}]

    def run_departure(
        *,
        input_excel_path: str,
        output_excel_path: str | None = None,
        progress_callback=None,
        **kwargs,
    ) -> dict:
        if departure_api == "none":
            raise RuntimeError(
                "departure scheduler module missing. "
                "請部署 `departure_duty_scheduler.py`（或舊版 `departure_scheduler.py`）。"
            )
        assert departure_mod is not None
        out_path = str(output_excel_path or "")
        if out_path.strip() == "":
            raise ValueError("output_excel_path is required for departure mode.")

        report_path = str(Path(out_path).with_name("departure_report.txt"))
        if departure_api == "legacy":
            run_fn = getattr(departure_mod, "run_departure_scheduler", None)
            if not callable(run_fn):
                raise RuntimeError("legacy departure module 缺少 run_departure_scheduler")
            result = run_fn(
                input_excel_path=input_excel_path,
                output_excel_path=out_path,
                search_best_roster=True,
                search_patience=5,
                require_all_pulls_nonzero=False,
                rescue_fill=True,
                debug=False,
                progress_callback=progress_callback,
                priority_mode=str(kwargs.get("priority_mode", "team1")),
                custom_order=str(kwargs.get("custom_order", "fairness,shift_count")),
                score_order=str(kwargs.get("score_order", "fairness,shift,pull")),
            )
            return {
                "tries": int(result.get("tries", 1) or 1),
                "best_score_100": float(result.get("best_score_100", 0.0) or 0.0),
                "chart_data": result.get("chart_data", {}) or {},
                "mode_used": str(result.get("mode_used", "legacy")),
                "status": str(result.get("status", "legacy_ok")),
                "total_shortage_slots": int(result.get("total_shortage_slots", 0) or 0),
            }

        settings = departure_mod.SolverSettings(
            weight_last_hour_work=50,
            weight_group_fairness=8,
            weight_target_deviation=3,
            weight_same_hour_consistency=12,
            weight_single_slot_fragment=18,
            weight_shortage_slot=100000,
            auto_gate_max_slots=6,
            max_consecutive_work_slots=6,
            early_max_work_slots=14,
            late_max_work_slots=15,
            enforce_shift_work_caps=False,
            weight_shift_cap_excess=30,
            feasibility_mode="hard",
            max_time_sec=30,
        )

        heartbeat_thread = None
        heartbeat_stop = threading.Event()
        # hard mode may fallback to allow_shortage, so reserve roughly 2x solve budget.
        expected_sec = max(20, int(settings.max_time_sec) * 2 + 10)

        if callable(progress_callback):
            try:
                progress_callback(1, 100)
            except Exception:
                pass

            def _heartbeat() -> None:
                start = time.time()
                while not heartbeat_stop.wait(1.0):
                    elapsed = max(0.0, time.time() - start)
                    cur = min(99, max(1, int((elapsed / expected_sec) * 100)))
                    try:
                        progress_callback(cur, 100)
                    except Exception:
                        return

            heartbeat_thread = threading.Thread(target=_heartbeat, daemon=True)
            heartbeat_thread.start()

        try:
            result = departure_mod.run_pipeline(
                input_path=input_excel_path,
                output_excel_path=out_path,
                report_path=report_path,
                settings=settings,
                dry_run=False,
                fallback_to_allow_shortage=False,
            )
        finally:
            if heartbeat_thread is not None:
                heartbeat_stop.set()
                heartbeat_thread.join(timeout=0.2)

        if callable(progress_callback):
            try:
                progress_callback(100, 100)
            except Exception:
                pass

        all_covered = bool(result.get("all_covered", False))
        total_shortage = int(result.get("total_shortage_slots", 0) or 0)
        if not all_covered:
            raise RuntimeError(
                "出境勤務表存在開天窗，已依硬規則視為不可接受。"
                f"status={result.get('status', '')}, shortage_slots={total_shortage}。"
                f"請查看診斷報告：{report_path}"
            )
        score = 100.0 if all_covered else max(0.0, 100.0 - float(total_shortage))
        departure_stats = _build_departure_chart_data(out_path)
        return {
            "tries": 1,
            "best_score_100": score,
            "chart_data": {"departure": departure_stats},
            "mode_used": result.get("mode_used", ""),
            "status": result.get("status", ""),
            "total_shortage_slots": total_shortage,
        }

    return {
        "monthly": {"run": run_monthly, "validate": validate_monthly},
        "departure": {"run": run_departure, "validate": validate_departure},
    }


app = FastAPI(title="Airport Scheduler MVP")
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# Simple in-memory store for recent results (token -> metadata)
_RESULTS: dict[str, dict] = {}
_RESULT_TTL_SEC = 60 * 10


def _store_result(
    token: str,
    out_path: Path,
    tmpdir: Path,
    tries: int,
    best_score_100: float,
    chart_data: dict,
    mode: str,
) -> None:
    _RESULTS[token] = {
        "out_path": out_path,
        "tmpdir": tmpdir,
        "tries": tries,
        "best_score_100": best_score_100,
        "chart_data": chart_data,
        "mode": _mode_key_or_default(mode),
        "ts": time.time(),
        "status": "done",
        "progress": 1.0,
    }


def _pop_result(token: str) -> dict | None:
    data = _RESULTS.pop(token, None)
    return data


def _get_result(token: str) -> dict | None:
    _cleanup_expired_results()
    data = _RESULTS.get(token)
    if data:
        data["ts"] = time.time()
    return data


def _esc(v) -> str:
    return html.escape(str(v if v is not None else ""), quote=True)


def _init_progress(token: str, tmpdir: Path, min_tries: int, mode: str) -> None:
    _RESULTS[token] = {
        "tmpdir": tmpdir,
        "ts": time.time(),
        "start_ts": time.time(),
        "status": "running",
        "progress": 0.0,
        "tries": 0,
        "min_tries": int(min_tries),
        "mode": _mode_key_or_default(mode),
    }


def _set_progress(token: str, current_try: int, max_tries: int) -> None:
    data = _RESULTS.get(token)
    if not data:
        return
    data["ts"] = time.time()
    total = max(1, int(max_tries))
    data["min_tries"] = total
    cur = max(0, int(current_try))
    data["tries"] = cur
    data["progress"] = min(1.0, cur / total)


def _set_error(token: str, msg: str) -> None:
    data = _RESULTS.get(token)
    if not data:
        return
    data["status"] = "error"
    data["message"] = msg


def _format_validation_errors(errors: list[dict]) -> str:
    lines = []
    for err in errors:
        sheet = str(err.get("sheet", "") or "")
        cols = err.get("columns", []) or []
        reason = str(err.get("reason", "") or "")
        cells = err.get("cells", []) or []
        col_text = ", ".join([str(c) for c in cols if str(c).strip() != ""])
        if reason and col_text:
            line = f"{sheet}: {reason} -> {col_text}"
        elif reason:
            line = f"{sheet}: {reason}"
        elif col_text:
            line = f"{sheet}: {col_text}"
        else:
            line = f"{sheet}: 欄位不符合模板"
        lines.append(line)
        if cells:
            for cell in cells:
                day = cell.get("day", None)
                person = str(cell.get("person", "") or "").strip()
                value = str(cell.get("value", "") or "").strip()
                if day is None or person == "":
                    continue
                if value != "":
                    lines.append(f"請檢查{int(day)}號的{person}（值={value}）")
                else:
                    lines.append(f"請檢查{int(day)}號的{person}")
    return "\n".join(lines) if lines else "輸入檔案格式不正確。"


def _hex_color(rgb) -> str:
    if not rgb:
        return ""
    try:
        v = rgb[-6:]
        return f"#{v}"
    except Exception:
        return ""


def _shift_window_to_group(shift_window: str) -> str:
    s = str(shift_window or "").strip()
    if "-" in s:
        start = s.split("-", 1)[0].strip()
    else:
        start = s
    try:
        h = int(str(start).split(":")[0])
    except Exception:
        return "Unknown"
    if h in (5, 6):
        return "Early"
    if h in (7, 8):
        return "Late"
    return "Unknown"


def _build_departure_chart_data(output_excel_path: str) -> dict:
    empty = {"work": {"Early": [], "Late": []}, "auto": {"Early": [], "Late": []}}
    try:
        summary = pd.read_excel(output_excel_path, sheet_name="Summary")
    except Exception:
        return empty

    req_cols = {"name", "shift_window", "worked_minutes", "auto_gate_minutes"}
    if not req_cols.issubset(set(summary.columns)):
        return empty

    work = {"Early": [], "Late": []}
    auto = {"Early": [], "Late": []}
    for _, row in summary.iterrows():
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        grp = _shift_window_to_group(str(row.get("shift_window", "")))
        if grp not in ("Early", "Late"):
            continue
        try:
            worked_h = float(row.get("worked_minutes", 0) or 0) / 60.0
        except Exception:
            worked_h = 0.0
        try:
            auto_h = float(row.get("auto_gate_minutes", 0) or 0) / 60.0
        except Exception:
            auto_h = 0.0
        has_auto_skill_raw = row.get("has_auto_gate_skill", 1)
        has_auto_skill = False
        if isinstance(has_auto_skill_raw, str):
            has_auto_skill = has_auto_skill_raw.strip().lower() in {"1", "true", "yes", "y"}
        else:
            try:
                has_auto_skill = bool(int(has_auto_skill_raw))
            except Exception:
                has_auto_skill = bool(has_auto_skill_raw)
        work[grp].append((name, round(worked_h, 1)))
        if has_auto_skill:
            auto[grp].append((name, round(auto_h, 1)))

    for grp in ("Early", "Late"):
        work[grp].sort(key=lambda x: x[1], reverse=True)
        auto[grp].sort(key=lambda x: x[1], reverse=True)
    return {"work": work, "auto": auto}


def _cleanup_expired_results() -> None:
    now = time.time()
    expired = [k for k, v in _RESULTS.items() if now - float(v.get("ts", 0)) > _RESULT_TTL_SEC]
    for k in expired:
        v = _RESULTS.pop(k, None)
        if not v:
            continue
        tmpdir = v.get("tmpdir")
        if isinstance(tmpdir, Path):
            shutil.rmtree(tmpdir, ignore_errors=True)


def _find_manual_preview_url() -> str | None:
    static_dir = BASE_DIR / "static"
    preferred = [
        static_dir / "manual.pdf",
        static_dir / "使用說明書.pdf",
    ]
    candidates: list[Path] = []
    for p in preferred:
        if p.exists() and p.is_file():
            candidates.append(p)

    if not candidates:
        candidates = [p for p in static_dir.glob("*.pdf") if p.is_file()]
    if not candidates:
        return None

    # Always use the newest file and append mtime to bypass browser cache.
    chosen = max(candidates, key=lambda p: p.stat().st_mtime)
    ver = int(chosen.stat().st_mtime)
    return f"/static/{chosen.name}?v={ver}"

    


@app.get("/", response_class=HTMLResponse)
def home():
    return RedirectResponse(url="/monthly", status_code=307)


def _render_scheduler_home(request: Request, default_mode: str = "monthly") -> HTMLResponse:
    manual_preview_url = _find_manual_preview_url()
    modes = [
        {
            "key": key,
            "label": mode["label"],
            "desc": mode["desc"],
        }
        for key, mode in _SCHEDULE_MODES.items()
    ]
    resp = templates.TemplateResponse(
        "home.html",
        {
            "request": request,
            "template_url": "https://drive.google.com/drive/folders/1mNXtRv5olbJQAGhnVy30mBoa8m4nTAJT?usp=sharing",
            "manual_url": manual_preview_url or "/static/manual.pdf",
            "manual_preview_url": manual_preview_url,
            "modes": modes,
            "default_mode": _mode_key_or_default(default_mode),
        },
    )
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.get("/monthly", response_class=HTMLResponse)
def monthly_home(request: Request):
    return _render_scheduler_home(request, default_mode="monthly")


@app.post("/run")
async def run(
    background_tasks: BackgroundTasks,
    request: Request,
    file: UploadFile = File(...),
    schedule_mode: str = Form("monthly"),
    priority_mode: str = Form("team1"),
    custom_order: str = Form("fairness,shift_count"),
    score_order: str = Form("fairness,shift,pull"),
):
    mode_key = _mode_key_or_default(schedule_mode)
    mode_cfg = _SCHEDULE_MODES.get(mode_key, _SCHEDULE_MODES["monthly"])
    # 1) 基本檢查：副檔名
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")
    # 2) 每個 request 使用獨立 temp 資料夾（避免互踩）
    tmpdir = Path(tempfile.mkdtemp(prefix="airport_web_"))
    in_path = tmpdir / "input.xlsx"
    out_path = tmpdir / "output.xlsx"

    try:
        _cleanup_expired_results()
        # 3) 存上傳檔到 temp
        try:
            with in_path.open("wb") as f:
                shutil.copyfileobj(file.file, f)
        finally:
            await file.close()

        # 4) 檔案大小限制（例如 10MB）
        if in_path.stat().st_size > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large (max 10MB).")

        # 4.1) 輸入資料檢查：立即回報錯誤欄位並返回首頁
        scheduler_funcs = _get_scheduler_funcs()
        mode_funcs = scheduler_funcs.get(mode_key, scheduler_funcs["monthly"])
        run_scheduler = mode_funcs["run"]
        validate_input_excel = mode_funcs["validate"]
        validation_errors = validate_input_excel(str(in_path))
        if validation_errors:
            msg = _format_validation_errors(validation_errors)
            shutil.rmtree(tmpdir, ignore_errors=True)
            return templates.TemplateResponse(
                "validation_error.html",
                {
                    "request": request,
                    "message": msg,
                },
                status_code=400,
            )

        token = uuid.uuid4().hex
        _init_progress(token, tmpdir, 100, mode_key)

        def _run_job() -> None:
            try:
                def _cb(cur: int, mx: int) -> None:
                    _set_progress(token, cur, mx)

                run_kwargs = mode_cfg.get("run_kwargs", {}) or {}
                result = run_scheduler(
                    input_excel_path=str(in_path),
                    output_excel_path=str(out_path),
                    debug=False,
                    progress_callback=_cb,
                    priority_mode=priority_mode,
                    custom_order=custom_order,
                    score_order=score_order,
                    **run_kwargs,
                )

                tries_used = int(result.get("tries", 0) or 0)
                best_score = float(result.get("best_score_100", 0.0) or 0.0)
                chart_data = result.get("chart_data", {}) or {}
                _store_result(token, out_path, tmpdir, tries_used, best_score, chart_data, mode_key)
            except Exception as e:
                _set_error(token, str(e))

        background_tasks.add_task(_run_job)

        return templates.TemplateResponse(
            "running.html",
            {
                "request": request,
                "token": token,
                "mode_label": mode_cfg["label"],
            },
        )

    except Exception:
        shutil.rmtree(tmpdir, ignore_errors=True)
        raise


@app.get("/download/{token}")
def download(token: str, background_tasks: BackgroundTasks):
    data = _get_result(token)
    if not data:
        raise HTTPException(status_code=404, detail="Result expired or not found.")

    out_path = data.get("out_path")
    tmpdir = data.get("tmpdir")
    if not isinstance(out_path, Path) or not out_path.exists():
        if isinstance(tmpdir, Path):
            shutil.rmtree(tmpdir, ignore_errors=True)
        raise HTTPException(status_code=404, detail="File missing.")

    mode_key = _mode_key_or_default(data.get("mode"))
    mode_suffix = "departure_duty" if mode_key == "departure" else "monthly_roster"
    return FileResponse(
        path=str(out_path),
        filename=f"{mode_suffix}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/progress/{token}")
def progress(token: str):
    data = _get_result(token)
    if not data:
        return JSONResponse({"status": "error", "message": "Result expired or not found."}, status_code=404)
    now = time.time()
    start_ts = float(data.get("start_ts", now) or now)
    tries = int(data.get("tries", 0) or 0)
    max_tries = int(data.get("min_tries", 100) or 100)
    eta_sec = None
    elapsed = max(0.0, now - start_ts)
    if tries > 0 and elapsed > 0.5:
        rate = tries / elapsed
        if rate > 0:
            remaining = max(0, max_tries - tries)
            eta_sec = remaining / rate
    return JSONResponse(
        {
            "status": data.get("status", "running"),
            "progress": float(data.get("progress", 0.0) or 0.0),
            "tries": tries,
            "max_tries": max_tries,
            "eta_sec": eta_sec,
            "message": data.get("message", ""),
        }
    )


@app.get("/preview/{token}", response_class=HTMLResponse)
def preview(token: str, request: Request):
    data = _get_result(token)
    if not data:
        raise HTTPException(status_code=404, detail="Result expired or not found.")

    out_path = data.get("out_path")
    if not isinstance(out_path, Path) or not out_path.exists():
        raise HTTPException(status_code=404, detail="File missing.")

    mode_key = _mode_key_or_default(data.get("mode"))
    mode_label = _SCHEDULE_MODES.get(mode_key, _SCHEDULE_MODES["monthly"])["label"]

    try:
        wb = load_workbook(out_path, data_only=True)
        if mode_key == "departure" and "Dispatch" in wb.sheetnames:
            ws = wb["Dispatch"]
        else:
            ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel: {e}")

    chart_data = data.get("chart_data", {}) or {}
    shift_charts = chart_data.get("shift", {}) or {}
    skill_charts = chart_data.get("skill", {}) or {}
    pull_charts = chart_data.get("pull", {}) or {}
    dep_charts = chart_data.get("departure", {}) or {}

    max_rows = ws.max_row
    max_cols = ws.max_column

    col_pct = 100 / max(1, max_cols)
    min_col_width = "96px" if mode_key == "departure" else "110px"
    # Build HTML table with basic fill colors + bold
    rows_html = []
    for r in range(1, max_rows + 1):
        cells = []
        for c in range(1, max_cols + 1):
            cell = ws.cell(row=r, column=c)
            val = _esc("" if cell.value is None else str(cell.value)).replace("\n", "<br/>")
            styles = []
            fill = cell.fill
            if fill and getattr(fill, "fill_type", None) not in (None, "none"):
                fg = getattr(fill, "fgColor", None)
                rgb = getattr(fg, "rgb", None) if fg else None
                color = _hex_color(rgb)
                if color:
                    styles.append(f"background-color: {color};")
            if cell.font and cell.font.bold:
                styles.append("font-weight: 700;")
            styles.append(f"width:{col_pct:.3f}%")
            styles.append(f"min-width:{min_col_width}")
            styles.append("vertical-align:top")
            style_attr = f' style="{" ".join(styles)}"'
            tag = "th" if r == 1 else "td"
            cells.append(f"<{tag}{style_attr}>{val}</{tag}>")
        rows_html.append("<tr>" + "".join(cells) + "</tr>")
    table_html = (
        "<table style='table-layout: fixed; width: 100%; border-collapse: collapse;' "
        "border='1' cellspacing='0' cellpadding='4'>"
        + "".join(rows_html)
        + "</table>"
    )

    def _render_dual_group_card(title: str, groups: dict) -> str:
        a_list = groups.get("A", []) or []
        b_list = groups.get("B", []) or []
        max_val = 0
        for _, v in a_list + b_list:
            try:
                max_val = max(max_val, int(v))
            except Exception:
                pass
        max_val = max(1, max_val)

        def _bars(items: list, tone_class: str) -> str:
            rows = []
            for name, v in items:
                try:
                    val = int(v)
                except Exception:
                    val = 0
                width_pct = int((val / max_val) * 100)
                safe_name = _esc(name)
                rows.append(
                    "<div class='viz-row'>"
                    f"<div class='viz-name'>{safe_name}</div>"
                    "<div class='viz-bar-track'>"
                    f"<div class='viz-bar {tone_class}' style='width:{width_pct}%;'></div>"
                    "</div>"
                    f"<div class='viz-value'>{val}</div>"
                    "</div>"
                )
            return "".join(rows) if rows else "<div class='viz-empty'>無資料</div>"

        return (
            "<div class='viz-card'>"
            f"<div class='viz-title'>{_esc(title)}</div>"
            "<div class='viz-grid'>"
            "<div class='viz-col viz-col-split'>"
            "<div class='viz-col-title'>A組</div>"
            f"{_bars(a_list, 'viz-tone-a')}"
            "</div>"
            "<div class='viz-col'>"
            "<div class='viz-col-title'>B組</div>"
            f"{_bars(b_list, 'viz-tone-b')}"
            "</div>"
            "</div>"
            "</div>"
        )

    charts_html = "".join(_render_dual_group_card(sh, groups) for sh, groups in shift_charts.items())
    skill_html = "".join(_render_dual_group_card(sk, groups) for sk, groups in skill_charts.items())
    pull_html = "".join(_render_dual_group_card(title, groups) for title, groups in pull_charts.items())

    def _fmt_hour_val(v) -> str:
        try:
            f = float(v)
        except Exception:
            return "0.0"
        if abs(f - round(f)) < 1e-9:
            return f"{int(round(f))}.0"
        return f"{f:.1f}"

    def _render_early_late_card(title: str, groups: dict) -> str:
        early_list = groups.get("Early", []) or []
        late_list = groups.get("Late", []) or []
        max_val = 0.0
        for _, v in early_list + late_list:
            try:
                max_val = max(max_val, float(v))
            except Exception:
                pass
        max_val = max(0.1, max_val)

        def _bars(items: list, tone_class: str) -> str:
            rows = []
            for name, v in items:
                try:
                    val = float(v)
                except Exception:
                    val = 0.0
                width_pct = int((val / max_val) * 100)
                safe_name = _esc(name)
                rows.append(
                    "<div class='viz-row'>"
                    f"<div class='viz-name'>{safe_name}</div>"
                    "<div class='viz-bar-track'>"
                    f"<div class='viz-bar {tone_class}' style='width:{width_pct}%;'></div>"
                    "</div>"
                    f"<div class='viz-value'>{_fmt_hour_val(val)}h</div>"
                    "</div>"
                )
            return "".join(rows) if rows else "<div class='viz-empty'>無資料</div>"

        return (
            "<div class='viz-card'>"
            f"<div class='viz-title'>{_esc(title)}</div>"
            "<div class='viz-grid'>"
            "<div class='viz-col viz-col-split'>"
            "<div class='viz-col-title'>早班（05/06）</div>"
            f"{_bars(early_list, 'viz-tone-a')}"
            "</div>"
            "<div class='viz-col'>"
            "<div class='viz-col-title'>晚班（07/08）</div>"
            f"{_bars(late_list, 'viz-tone-b')}"
            "</div>"
            "</div>"
            "</div>"
        )

    departure_work_html = ""
    departure_auto_html = ""
    if dep_charts:
        departure_work_html = _render_early_late_card("總上勤時數（小時）", dep_charts.get("work", {}) or {})
        departure_auto_html = _render_early_late_card("自動通關時數（小時）", dep_charts.get("auto", {}) or {})

    return templates.TemplateResponse(
        "preview.html",
        {
            "request": request,
            "token": token,
            "mode_label": mode_label,
            "table_html": table_html,
            "charts_html": charts_html,
            "skill_html": skill_html,
            "pull_html": pull_html,
            "departure_work_html": departure_work_html,
            "departure_auto_html": departure_auto_html,
        },
    )


@app.get("/report/{token}")
def report_pdf(token: str):
    data = _get_result(token)
    if not data:
        raise HTTPException(status_code=404, detail="Result expired or not found.")

    chart_data = data.get("chart_data", {}) or {}
    shift_charts = chart_data.get("shift", {}) or {}
    skill_charts = chart_data.get("skill", {}) or {}
    pull_charts = chart_data.get("pull", {}) or {}
    dep_charts = chart_data.get("departure", {}) or {}

    # Register Chinese font
    font_path = BASE_DIR / "fonts" / "static" / "NotoSansTC-Regular.ttf"
    if font_path.exists():
        pdfmetrics.registerFont(TTFont("NotoSansTC", str(font_path)))
        base_font = "NotoSansTC"
    else:
        base_font = "Helvetica"

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    def _draw_chart(title: str, groups: dict, y: float, left_label: str = "A", right_label: str = "B") -> float:
        c.setFont(base_font, 10)
        c.drawString(20 * mm, y, title)
        y -= 6 * mm

        a_list = groups.get("A", []) or []
        b_list = groups.get("B", []) or []
        max_val = 1.0
        for _, v in a_list + b_list:
            try:
                max_val = max(max_val, float(v))
            except Exception:
                pass

        def _draw_group(label: str, items: list, color: str, x0: float, y0: float) -> float:
            c.setFont(base_font, 8)
            c.drawString(x0, y0, label)
            y1 = y0 - 4 * mm
            for name, v in items:
                try:
                    val = float(v)
                except Exception:
                    val = 0.0
                bar_w = 60 * mm * (val / max_val)
                c.setFillColor(HexColor(color))
                c.rect(x0 + 22 * mm, y1 - 2, bar_w, 3 * mm, stroke=0, fill=1)
                c.setFillColor(HexColor("#000000"))
                c.drawString(x0, y1, str(name)[:10])
                if abs(val - round(val)) < 1e-9:
                    val_text = f"{int(round(val))}"
                else:
                    val_text = f"{val:.1f}"
                c.drawRightString(x0 + 20 * mm, y1, val_text)
                y1 -= 4 * mm
                if y1 < 20 * mm:
                    c.showPage()
                    y1 = height - 20 * mm
            return y1

        left_x = 20 * mm
        right_x = width / 2 + 5 * mm
        y_left = _draw_group(left_label, a_list, "#F9E27D", left_x, y)
        y_right = _draw_group(right_label, b_list, "#9AD59A", right_x, y)
        y = min(y_left, y_right) - 8 * mm
        if y < 25 * mm:
            c.showPage()
            y = height - 20 * mm
        return y

    y = height - 20 * mm
    for title, groups in shift_charts.items():
        y = _draw_chart(title, groups, y)
    for title, groups in skill_charts.items():
        y = _draw_chart(title, groups, y)
    for title, groups in pull_charts.items():
        y = _draw_chart(title, groups, y)
    if dep_charts:
        work_groups = {
            "A": (dep_charts.get("work", {}) or {}).get("Early", []) or [],
            "B": (dep_charts.get("work", {}) or {}).get("Late", []) or [],
        }
        auto_groups = {
            "A": (dep_charts.get("auto", {}) or {}).get("Early", []) or [],
            "B": (dep_charts.get("auto", {}) or {}).get("Late", []) or [],
        }
        y = _draw_chart("總上勤時數（小時）", work_groups, y, left_label="早班", right_label="晚班")
        y = _draw_chart("自動通關時數（小時）", auto_groups, y, left_label="早班", right_label="晚班")

    c.save()
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=report.pdf"})


if __name__ == "__main__":
    import os
    import uvicorn

    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    uvicorn.run(app, host=host, port=port)
